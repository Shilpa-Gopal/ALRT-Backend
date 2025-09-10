from flask import Flask, jsonify, request, abort, send_file
import os
from werkzeug.utils import secure_filename
from flask_cors import CORS
import io
import xlsxwriter
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np
from werkzeug.security import generate_password_hash, check_password_hash
from app import create_app, db
from app.models import User, Project, Citation
from app.ml_system import LiteratureReviewSystem
import hashlib
import re
import requests
from datetime import datetime, timezone
from sqlalchemy import text

app = create_app()
MIN_LABELS_PER_CATEGORY = 5
MAX_LABELS_PER_CATEGORY = 10

import json
import numpy as np

def make_json_serializable(obj):
    """Convert numpy types to JSON serializable Python types"""
    if isinstance(obj, dict):
        return {k: make_json_serializable(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [make_json_serializable(item) for item in obj]
    elif isinstance(obj, np.integer):
        return int(obj)
    elif isinstance(obj, np.floating):
        return float(obj)
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    elif pd.isna(obj):
        return None
    else:
        return obj

@app.route('/api/db-test', methods=['GET'])
def test_database():
    try:
        with app.app_context():
            result = db.session.execute(text('SELECT version();'))
            version = result.fetchone()[0]
            return jsonify({
                "status": "success",
                "message": "Database connection successful",
                "postgresql_version": version
            }), 200
    except Exception as e:
        return jsonify({
            "status": "error", 
            "message": str(e)
        }), 500

def normalize_text(text):
    """Normalize text for better matching"""
    if pd.isna(text):
        return ""
    text = str(text).lower()
    # Remove extra spaces, punctuation, and normalize
    text = re.sub(r'[^\w\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def create_text_hash(title, abstract, authors=None):
    """Create a hash for quick duplicate detection"""
    combined = f"{normalize_text(title)} {normalize_text(abstract)}"
    if authors:
        combined += f" {normalize_text(authors)}"
    return hashlib.md5(combined.encode()).hexdigest()

def find_year_column(df):
    """Find year-related column in dataframe (case insensitive)"""
    year_patterns = ['year', 'publication_year', 'pub_year', 'published_year', 
                    'publication_date', 'pub_date', 'date', 'published']

    for col in df.columns:
        for pattern in year_patterns:
            if pattern.lower() in col.lower():
                return col
    return None

def extract_year_from_value(value):
    """Extract year from various date formats"""
    if pd.isna(value):
        return None

    value_str = str(value)
    year_match = re.search(r'\b(19|20)\d{2}\b', value_str)
    if year_match:
        return int(year_match.group())
    return None

def calculate_completeness_score(row):
    """Calculate completeness score based on filled fields and content length"""
    score = 0
    total_fields = len(row)
    filled_fields = sum(1 for val in row if pd.notna(val) and str(val).strip())

    # Base score from field completion
    score += (filled_fields / total_fields) * 50

    # Additional score for abstract length
    if pd.notna(row.get('abstract')):
        abstract_len = len(str(row['abstract']))
        score += min(abstract_len / 10, 50)

    return score

def apply_keyword_filtering(citations, project):
    """Apply keyword filtering with prioritization for include keywords"""
    if not project.keywords:
        return citations

    include_keywords = project.keywords.get('include', [])
    exclude_keywords = project.keywords.get('exclude', [])

    filtered_citations = []
    for citation in citations:
        text = f"{citation.title} {citation.abstract}".lower()

        # First check exclude keywords
        should_exclude = False
        for exclude_kw in exclude_keywords:
            word = exclude_kw['word'].lower()
            frequency = exclude_kw.get('frequency', 1)
            occurrences = text.count(word)

            if occurrences >= frequency:
                should_exclude = True
                break

        if not should_exclude:
            # Calculate include keyword score for prioritization
            include_score = 0
            for include_word in include_keywords:
                word = include_word.lower()
                occurrences = text.count(word)
                include_score += occurrences

            # Add include score as metadata for later use
            citation._include_score = include_score
            filtered_citations.append(citation)

    # Sort by include keyword score (highest first) to prioritize include keywords
    filtered_citations.sort(key=lambda c: getattr(c, '_include_score', 0), reverse=True)

    return filtered_citations

def process_duplicates_and_create_citations(df, project_id, current_iteration):
    """Enhanced duplicate processing with column standardization and new detection logic"""

    app.logger.info(f"Processing {len(df)} citations for enhanced duplicates")

    # Normalize column names and standardize to consistent format
    df_normalized = standardize_dataframe_columns(df.copy())

    # Check required columns
    if 'title' not in df_normalized.columns or 'abstract' not in df_normalized.columns:
        return {
            'error': f"File must contain 'title' and 'abstract' columns. Found: {list(df.columns)}",
            'citations': [],
            'duplicates_removed': 0,
            'removal_strategy': '',
            'duplicate_details': []
        }

    # Check if we have author and year information
    has_authors = 'authors' in df_normalized.columns and not df_normalized['authors'].isna().all()
    has_year = 'publication_year' in df_normalized.columns and not df_normalized['publication_year'].isna().all()

    app.logger.info(f"Enhanced duplicate detection: Authors present: {has_authors}, Year present: {has_year}")
    app.logger.info(f"DataFrame columns after standardization: {list(df_normalized.columns)}")
    app.logger.info(f"DataFrame shape: {df_normalized.shape}")
    app.logger.info(f"Sample data - First row title: '{df_normalized.iloc[0]['title'] if len(df_normalized) > 0 else 'N/A'}'")
    app.logger.info(f"Sample data - First row abstract: '{str(df_normalized.iloc[0]['abstract'])[:50] if len(df_normalized) > 0 else 'N/A'}...'")

    # Apply the new enhanced duplicate processing logic
    result = process_enhanced_duplicates(df_normalized, has_authors, has_year)

    if result['error']:
        return result

    # Create Citation objects from the kept citations
    new_citations = []
    for _, row in enumerate(result['citations_to_keep']):
        citation = Citation(
            title=str(row['title']),
            abstract=str(row['abstract']),
            project_id=project_id,
            iteration=current_iteration
        )
        new_citations.append(citation)

    duplicates_removed = len(df) - len(new_citations)
    app.logger.info(f"Enhanced duplicate removal completed: {duplicates_removed} duplicates removed, keeping {len(new_citations)} citations")
    app.logger.info(f"DEBUG: Original DataFrame length: {len(df)}, Citations kept: {len(new_citations)}")
    app.logger.info(f"DEBUG: Duplicate details: {result['duplicate_details']}")

    return {
        'error': None,
        'citations': new_citations,
        'duplicates_removed': duplicates_removed,
        'removal_strategy': result['strategy'],
        'duplicate_details': result['duplicate_details'][:10]  # Return first 10 for response
    }

def select_best_citation(group, year_col, duplicate_details):
    """Select the best citation from a group of duplicates"""
    if len(group) == 1:
        return group.iloc[0]

    # Strategy 1: Use year if available
    if year_col:
        group['year_extracted'] = group[year_col].apply(extract_year_from_value)
        valid_years = group[group['year_extracted'].notna()]

        if len(valid_years) > 0:
            # Keep the most recent
            best = valid_years.loc[valid_years['year_extracted'].idxmax()]

            # Log the duplicate removal
            for _, other in group.iterrows():
                if other.name != best.name:
                    duplicate_details.append({
                        'kept': f"{best['title'][:50]}...",
                        'removed': f"{other['title'][:50]}...",
                        'reason': f'Year-based selection'
                    })

            return best

    # Strategy 2: Use completeness score
    group['completeness'] = group.apply(calculate_completeness_score, axis=1)
    best = group.loc[group['completeness'].idxmax()]

    # Log the duplicate removal
    for _, other in group.iterrows():
        if other.name != best.name:
            duplicate_details.append({
                'kept': f"{best['title'][:50]}...",
                'removed': f"{other['title'][:50]}...",
                'reason': f'Completeness-based selection'
            })

    return best

def remove_similar_citations(citations_list, authors_col, duplicate_details):
    """Use TF-IDF to find and remove similar citations"""
    if len(citations_list) < 2:
        return citations_list

    # Prepare text for TF-IDF
    texts = []
    for citation in citations_list:
        text = f"{normalize_text(citation['title'])} {normalize_text(citation['abstract'])}"
        if authors_col and pd.notna(citation.get(authors_col)):
            text += f" {normalize_text(citation[authors_col])}"
        texts.append(text)

    # Calculate TF-IDF similarity only if we have enough text variation
    try:
        vectorizer = TfidfVectorizer(max_features=500, stop_words='english', ngram_range=(1, 2))
        tfidf_matrix = vectorizer.fit_transform(texts)
        similarity_matrix = cosine_similarity(tfidf_matrix)

        # Find similar pairs (above 70% similarity)
        to_remove = set()
        for i in range(len(similarity_matrix)):
            for j in range(i + 1, len(similarity_matrix)):
                if similarity_matrix[i][j] > 0.7:
                    # Keep the one with higher completeness
                    citation_i = citations_list[i]
                    citation_j = citations_list[j]

                    score_i = calculate_completeness_score(citation_i)
                    score_j = calculate_completeness_score(citation_j)

                    if score_i >= score_j:
                        to_remove.add(j)
                        duplicate_details.append({
                            'kept': f"{citation_i['title'][:50]}...",
                            'removed': f"{citation_j['title'][:50]}...",
                            'reason': f'TF-IDF similarity: {similarity_matrix[i][j]:.2f}'
                        })
                    else:
                        to_remove.add(i)
                        duplicate_details.append({
                            'kept': f"{citation_j['title'][:50]}...",
                            'removed': f"{citation_i['title'][:50]}...",
                            'reason': f'TF-IDF similarity: {similarity_matrix[i][j]:.2f}'
                        })

        # Return citations not in removal set
        return [citation for i, citation in enumerate(citations_list) if i not in to_remove]

    except Exception as e:
        app.logger.warning(f"TF-IDF similarity calculation failed: {e}")
        return citations_list
    

def create_enhanced_duplicate_detail(kept_citation, removed_citation, reason, similarity_score=None):
    """Create enhanced duplicate detail with full information"""
    
    def convert_numpy_types(value):
        if pd.isna(value):
            return None
        if hasattr(value, 'item'):  # numpy scalar
            return value.item()
        return value
    
    def get_val(citation, key):
        return convert_numpy_types(citation[key]) if key in citation else None
    
    detail = {
        'kept': {
            'id': int(convert_numpy_types(kept_citation['id'])),
            'title': str(kept_citation['title']),
            'abstract': (str(kept_citation['abstract'])[:200] + '...'
                         if len(str(kept_citation['abstract'])) > 200
                         else str(kept_citation['abstract'])),
            'is_relevant': get_val(kept_citation, 'is_relevant'),
            'iteration': int(get_val(kept_citation, 'iteration') or 0)
        },
        'removed': {
            'id': int(convert_numpy_types(removed_citation['id'])),
            'title': str(removed_citation['title']),
            'abstract': (str(removed_citation['abstract'])[:200] + '...'
                         if len(str(removed_citation['abstract'])) > 200
                         else str(removed_citation['abstract'])),
            'is_relevant': get_val(removed_citation, 'is_relevant'),
            'iteration': int(get_val(removed_citation, 'iteration') or 0)
        },
        'reason': str(reason),
        'similarity_score': float(similarity_score) if similarity_score is not None else None,
        'detection_method': 'tfidf' if similarity_score is not None else 'exact_match',
        'timestamp': datetime.now(timezone.utc).isoformat()
    }
    
    return detail


CORS(app, resources={r"/api/*": {
    "origins": "*",
    "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    "allow_headers": ["Content-Type", "X-User-Id"]
}})

with app.app_context():
    db.create_all()

    # Migration: Add is_admin column to existing users if it doesn't exist
    try:
        # Check if is_admin column exists using PostgreSQL system tables
        from sqlalchemy import text
        result = db.session.execute(text("""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_name = 'user' AND column_name = 'is_admin'
        """)).fetchall()

        if not result:
            app.logger.info("Adding is_admin column to user table")
            db.session.execute(text("ALTER TABLE \"user\" ADD COLUMN is_admin BOOLEAN DEFAULT FALSE"))
            db.session.commit()
            app.logger.info("Migration completed: is_admin column added")
        else:
            app.logger.info("is_admin column already exists")
    except Exception as e:
        app.logger.error(f"Migration error: {str(e)}")
        db.session.rollback()

    # Migration: Add is_duplicate column to existing citations if it doesn't exist
    try:
        result = db.session.execute(text("""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_name = 'citation' AND column_name = 'is_duplicate'
        """)).fetchall()

        if not result:
            app.logger.info("Adding is_duplicate column to citation table")
            db.session.execute(text("ALTER TABLE citation ADD COLUMN is_duplicate BOOLEAN DEFAULT FALSE"))
            db.session.commit()
            app.logger.info("Migration completed: is_duplicate column added")
        else:
            app.logger.info("is_duplicate column already exists")
    except Exception as e:
        app.logger.error(f"Migration error for is_duplicate: {str(e)}")
        db.session.rollback()
    
    # Migration: Add new project status fields
    try:
        # Check if new columns exist
        result = db.session.execute(text("""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_name = 'project' AND column_name IN 
            ('duplicates_removed', 'duplicates_count', 'keywords_selected', 'citations_count')
        """)).fetchall()
        
        existing_columns = [row[0] for row in result]
        
        # Add missing columns
        if 'duplicates_removed' not in existing_columns:
            app.logger.info("Adding duplicates_removed column to project table")
            db.session.execute(text("ALTER TABLE project ADD COLUMN duplicates_removed BOOLEAN DEFAULT FALSE"))
        
        if 'duplicates_count' not in existing_columns:
            app.logger.info("Adding duplicates_count column to project table")
            db.session.execute(text("ALTER TABLE project ADD COLUMN duplicates_count INTEGER DEFAULT 0"))
        
        if 'keywords_selected' not in existing_columns:
            app.logger.info("Adding keywords_selected column to project table")
            db.session.execute(text("ALTER TABLE project ADD COLUMN keywords_selected BOOLEAN DEFAULT FALSE"))
        
        if 'citations_count' not in existing_columns:
            app.logger.info("Adding citations_count column to project table")
            db.session.execute(text("ALTER TABLE project ADD COLUMN citations_count INTEGER DEFAULT 0"))
        
        db.session.commit()
        app.logger.info("Migration completed: Project status fields added")
        
    except Exception as e:
        app.logger.error(f"Migration error for project status fields: {str(e)}")
        db.session.rollback()
        
    # Migration: Add duplicate details storage fields
    try:
        # Check if new duplicate storage columns exist
        result = db.session.execute(text("""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_name = 'project' AND column_name IN 
            ('duplicate_details', 'processing_summary', 'removal_strategy')
        """)).fetchall()
        
        existing_columns = [row[0] for row in result]
        
        # Add missing columns
        if 'duplicate_details' not in existing_columns:
            app.logger.info("Adding duplicate_details column to project table")
            db.session.execute(text("ALTER TABLE project ADD COLUMN duplicate_details JSON DEFAULT '[]'"))
        
        if 'processing_summary' not in existing_columns:
            app.logger.info("Adding processing_summary column to project table")
            db.session.execute(text("ALTER TABLE project ADD COLUMN processing_summary JSON DEFAULT '{}'"))
        
        if 'removal_strategy' not in existing_columns:
            app.logger.info("Adding removal_strategy column to project table")
            db.session.execute(text("ALTER TABLE project ADD COLUMN removal_strategy VARCHAR(500)"))
        
        db.session.commit()
        app.logger.info("Migration completed: Duplicate details storage fields added")
        
    except Exception as e:
        app.logger.error(f"Migration error for duplicate details storage: {str(e)}")
        db.session.rollback()

    # Migration: Add enhanced duplicate detection fields
    try:
        # Check if new enhanced duplicate detection columns exist
        result = db.session.execute(text("""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_name = 'project' AND column_name IN 
            ('detection_method_used', 'year_resolution_count', 'abstract_resolution_count', 'columns_standardized')
        """)).fetchall()
        
        existing_columns = [row[0] for row in result]
        
        # Add missing columns
        if 'detection_method_used' not in existing_columns:
            app.logger.info("Adding detection_method_used column to project table")
            db.session.execute(text("ALTER TABLE project ADD COLUMN detection_method_used VARCHAR(100)"))
        
        if 'year_resolution_count' not in existing_columns:
            app.logger.info("Adding year_resolution_count column to project table")
            db.session.execute(text("ALTER TABLE project ADD COLUMN year_resolution_count INTEGER DEFAULT 0"))
        
        if 'abstract_resolution_count' not in existing_columns:
            app.logger.info("Adding abstract_resolution_count column to project table")
            db.session.execute(text("ALTER TABLE project ADD COLUMN abstract_resolution_count INTEGER DEFAULT 0"))
        
        if 'columns_standardized' not in existing_columns:
            app.logger.info("Adding columns_standardized column to project table")
            db.session.execute(text("ALTER TABLE project ADD COLUMN columns_standardized BOOLEAN DEFAULT FALSE"))
        
        db.session.commit()
        app.logger.info("Migration completed: Enhanced duplicate detection fields added")
        
    except Exception as e:
        app.logger.error(f"Migration error for enhanced duplicate detection: {str(e)}")
        db.session.rollback()

    # Migration: Add custom duplicate configuration fields
    try:
        result = db.session.execute(text("""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_name = 'project' AND column_name IN 
            ('duplicate_config', 'available_columns')
        """)).fetchall()
        existing_columns = [row[0] for row in result]

        if 'duplicate_config' not in existing_columns:
            app.logger.info("Adding duplicate_config column to project table")
            # Add without a default to avoid JSON parsing issues; app will handle None
            db.session.execute(text("ALTER TABLE project ADD COLUMN duplicate_config JSON"))

        if 'available_columns' not in existing_columns:
            app.logger.info("Adding available_columns column to project table")
            db.session.execute(text("ALTER TABLE project ADD COLUMN available_columns JSON DEFAULT '[]'"))

        db.session.commit()
        app.logger.info("Migration completed: Custom duplicate configuration fields added")
    except Exception as e:
        app.logger.error(f"Migration error for custom duplicate config: {str(e)}")
        db.session.rollback()

    # Migration: Add citation.metadata JSON column
    try:
        result = db.session.execute(text("""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_name = 'citation' AND column_name = 'metadata'
        """
        )).fetchall()
        if not result:
            app.logger.info("Adding metadata column to citation table")
            db.session.execute(text("ALTER TABLE citation ADD COLUMN metadata JSON DEFAULT '{}'"))
            db.session.commit()
            app.logger.info("Migration completed: metadata column added to citation")
    except Exception as e:
        app.logger.error(f"Migration error for citation.metadata: {str(e)}")
        db.session.rollback()




@app.route('/', methods=['GET'])
def home():
    return 'Backend API is running', 200


@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    try:
        # Test database connection
        db.session.execute(text('SELECT 1'))
        return jsonify({
            "status": "healthy",
            "database": "connected",
            "timestamp": datetime.utcnow().isoformat()
        }), 200
    except Exception as e:
        return jsonify({
            "status": "unhealthy",
            "database": "disconnected",
            "error": str(e),
            "timestamp": datetime.utcnow().isoformat()
        }), 503


@app.route('/api', methods=['GET'])
def api_info():
    """API information endpoint"""
    return jsonify({
        "name": "Literature Review API",
        "version": "1.0",
        "status": "running"
    }), 200


@app.route('/api/auth/signup', methods=['POST'])
def signup():
    data = request.get_json()

    if not all(k in data
               for k in ["first_name", "last_name", "email", "password"]):
        return jsonify({"error": "Missing required fields"}), 400

    if User.query.filter_by(email=data['email']).first():
        return jsonify({"error": "Email already registered"}), 400

    user = User(first_name=data['first_name'],
                last_name=data['last_name'],
                email=data['email'],
                password=generate_password_hash(data['password']))

    db.session.add(user)
    db.session.commit()

    return jsonify({"message": "User created successfully"}), 201


@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.get_json()

    if not all(k in data for k in ["email", "password"]):
        return jsonify({"error": "Missing email or password"}), 400

    user = User.query.filter_by(email=data['email']).first()

    if not user or not check_password_hash(user.password, data['password']):
        return jsonify({"error": "Invalid email or password"}), 401

    return jsonify({
        "user": {
            "id": user.id,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "email": user.email,
            "is_admin": user.is_admin
        }
    })


def require_admin():
    """Decorator to require admin access"""
    def decorator(f):
        def wrapper(*args, **kwargs):
            user_id = request.headers.get('X-User-Id')
            if not user_id:
                return jsonify({"error": "Unauthorized"}), 401

            user = User.query.get(int(user_id))
            if not user or not user.is_admin:
                return jsonify({"error": "Admin access required"}), 403

            return f(*args, **kwargs)
        wrapper.__name__ = f.__name__
        return wrapper
    return decorator


@app.route('/api/admin/users', methods=['GET'])
@require_admin()
def get_all_users():
    """Get all users with their projects - Admin only"""
    try:
        users = User.query.all()
        users_data = []

        for user in users:
            user_projects = []
            for project in user.projects:
                citations_count = Citation.query.filter_by(project_id=project.id).count()
                labeled_count = Citation.query.filter(
                    Citation.project_id == project.id,
                    Citation.is_relevant.isnot(None)
                ).count()

                user_projects.append({
                    "id": project.id,
                    "name": project.name,
                    "created_at": project.created_at,
                    "current_iteration": project.current_iteration,
                    "citations_count": citations_count,
                    "labeled_count": labeled_count,
                    "keywords": project.keywords,
                    "model_metrics": project.model_metrics
                })

            users_data.append({
                "id": user.id,
                "first_name": user.first_name,
                "last_name": user.last_name,
                "email": user.email,
                "password_hash": user.password,  # Include password hash for admin access
                "is_admin": user.is_admin,
                "projects": user_projects,
                "total_projects": len(user_projects)
            })

        return jsonify({
            "users": users_data,
            "total_users": len(users_data)
        })

    except Exception as e:
        app.logger.error(f"Error fetching users: {str(e)}")
        return jsonify({"error": "Failed to fetch users"}), 500


@app.route('/api/admin/users/<int:user_id>', methods=['GET'])
@require_admin()
def get_user_details(user_id):
    """Get detailed user information - Admin only"""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404

        projects = []
        for project in user.projects:
            citations = Citation.query.filter_by(project_id=project.id).all()
            citations_data = [{
                "id": c.id,
                "title": c.title,
                "abstract": c.abstract,
                "is_relevant": c.is_relevant,
                "iteration": c.iteration
            } for c in citations]

            projects.append({
                "id": project.id,
                "name": project.name,
                "created_at": project.created_at,
                "current_iteration": project.current_iteration,
                "keywords": project.keywords,
                "model_metrics": project.model_metrics,
                "citations": citations_data
            })

        return jsonify({
            "user": {
                "id": user.id,
                "first_name": user.first_name,
                "last_name": user.last_name,
                "email": user.email,
                "password_hash": user.password,
                "is_admin": user.is_admin,
                "projects": projects
            }
        })

    except Exception as e:
        app.logger.error(f"Error fetching user details: {str(e)}")
        return jsonify({"error": "Failed to fetch user details"}), 500


@app.route('/api/admin/users/<int:user_id>', methods=['PUT'])
@require_admin()
def update_user(user_id):
    """Update user information - Admin only"""
    try:
        data = request.get_json()
        user = User.query.get(user_id)

        if not user:
            return jsonify({"error": "User not found"}), 404

        # Update user fields
        if 'first_name' in data:
            user.first_name = data['first_name']
        if 'last_name' in data:
            user.last_name = data['last_name']
        if 'email' in data:
            # Check if email already exists for another user
            existing_user = User.query.filter_by(email=data['email']).first()
            if existing_user and existing_user.id != user_id:
                return jsonify({"error": "Email already exists"}), 400
            user.email = data['email']
        if 'password' in data:
            user.password = generate_password_hash(data['password'])
        if 'is_admin' in data:
            user.is_admin = data['is_admin']

        db.session.commit()

        return jsonify({
            "message": "User updated successfully",
            "user": {
                "id": user.id,
                "first_name": user.first_name,
                "last_name": user.last_name,
                "email": user.email,
                "is_admin": user.is_admin
            }
        })

    except Exception as e:
        app.logger.error(f"Error updating user: {str(e)}")
        db.session.rollback()
        return jsonify({"error": "Failed to update user"}), 500


@app.route('/api/admin/users/<int:user_id>/reset-password', methods=['POST'])
@require_admin()
def admin_reset_password(user_id):
    """Admin can reset any user's password"""
    try:
        # Handle both JSON and form data
        if request.is_json:
            data = request.get_json()
        else:
            data = request.form.to_dict()

        if not data:
            return jsonify({"error": "No data provided"}), 400

        user = User.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404

        new_password = data.get('new_password')
        if not new_password:
            return jsonify({"error": "New password is required"}), 400
        
        if len(new_password) < 6:
            return jsonify({"error": "Password must be at least 6 characters"}), 400

        user.password = generate_password_hash(new_password)
        db.session.commit()

        app.logger.info(f"Admin reset password for user {user.email}")

        return jsonify({
            "message": f"Password reset successfully for {user.email}",
            "new_password": new_password  # Only show in response for admin convenience
        })

    except Exception as e:
        app.logger.error(f"Error in admin password reset: {str(e)}")
        db.session.rollback()
        return jsonify({"error": "Failed to reset password", "details": str(e)}), 500


@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@require_admin()
def delete_user(user_id):
    """Delete user and all their projects - Admin only"""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Delete all citations for user's projects
        for project in user.projects:
            Citation.query.filter_by(project_id=project.id).delete()

        # Delete all user's projects
        Project.query.filter_by(user_id=user_id).delete()

        # Delete the user
        db.session.delete(user)
        db.session.commit()

        app.logger.info(f"Admin deleted user {user_id} and all associated data")
        return jsonify({"message": "User and all associated data deleted successfully"})

    except Exception as e:
        app.logger.error(f"Error deleting user: {str(e)}")
        db.session.rollback()
        return jsonify({"error": "Failed to delete user"}), 500


@app.route('/api/admin/projects', methods=['GET'])
@require_admin()
def get_all_projects():
    """Get all projects from all users - Admin only"""
    try:
        projects = Project.query.all()
        projects_data = []

        for project in projects:
            citations_count = Citation.query.filter_by(project_id=project.id).count()
            labeled_count = Citation.query.filter(
                Citation.project_id == project.id,
                Citation.is_relevant.isnot(None)
            ).count()

            projects_data.append({
                "id": project.id,
                "name": project.name,
                "created_at": project.created_at,
                "current_iteration": project.current_iteration,
                "user_id": project.user_id,
                "user_name": f"{project.user.first_name} {project.user.last_name}",
                "user_email": project.user.email,
                "citations_count": citations_count,
                "labeled_count": labeled_count,
                "keywords": project.keywords,
                "model_metrics": project.model_metrics
            })

        return jsonify({
            "projects": projects_data,
            "total_projects": len(projects_data)
        })

    except Exception as e:
        app.logger.error(f"Error fetching all projects: {str(e)}")
        return jsonify({"error": "Failed to fetch projects"}), 500


@app.route('/api/admin/projects/<int:project_id>', methods=['GET'])
@require_admin()
def get_project_details_admin(project_id):
    """Get detailed project information - Admin only"""
    try:
        user_id = request.headers.get('X-User-Id')
        if not user_id:
            return jsonify({"error": "Unauthorized"}), 401

        user_id_int = int(user_id)
        user = User.query.get(user_id_int)
        if not user:
            return jsonify({"error": "User not found"}), 401

        # Allow admin access to any project, regular users only their own
        if user.is_admin:
            project = Project.query.get(project_id)
        else:
            project = Project.query.filter_by(id=project_id, user_id=user_id_int).first()

        if not project:
            return jsonify({"error": "Project not found"}), 404

        citations = Citation.query.filter_by(project_id=project_id).all()
        labeled_count = Citation.query.filter(
            Citation.project_id == project_id,
            Citation.is_relevant.isnot(None)
        ).count()

        # Apply keyword filtering to citations
        non_duplicate_citations = [c for c in citations if not getattr(c, 'is_duplicate', False)]
        filtered_citations = apply_keyword_filtering(non_duplicate_citations, project)

        # Count labeled citations from filtered set
        filtered_labeled_count = sum(1 for c in filtered_citations if c.is_relevant is not None)

        return jsonify({
            "project": {
                "id": project.id,
                "name": project.name,
                "created_at": project.created_at,
                "current_iteration": project.current_iteration,
                "keywords": project.keywords,
                "model_metrics": project.model_metrics,
                "citations_count": len(filtered_citations),
                "labeled_count": filtered_labeled_count,
                "total_uploaded": len(citations),
                "duplicates_count": sum(1 for c in citations if getattr(c, 'is_duplicate', False)),
                "keyword_filtered_count": len(non_duplicate_citations) - len(filtered_citations),
                
                # New status fields
                "duplicates_removed": project.duplicates_removed,
                "keywords_selected": project.keywords_selected,
                "citations_count_stored": project.citations_count,
                "duplicates_count_stored": project.duplicates_count
            },
            "citations": [{
                "id": c.id,
                "title": c.title,
                "abstract": c.abstract,
                "is_relevant": c.is_relevant,
                "iteration": c.iteration,
                "is_duplicate": getattr(c, 'is_duplicate', False)
            } for c in filtered_citations]
        })

    except Exception as e:
        app.logger.error(f"Error fetching project details: {str(e)}")
        return jsonify({"error": "Failed to fetch project details"}), 500


@app.route('/api/admin/projects/<int:project_id>', methods=['PUT'])
@require_admin()
def update_project_admin(project_id):
    """Update project information - Admin only"""
    try:
        data = request.get_json()
        project = Project.query.get(project_id)

        if not project:
            return jsonify({"error": "Project not found"}), 404

        # Update project fields
        if 'name' in data:
            project.name = data['name']
        if 'keywords' in data:
            project.keywords = data['keywords']
        if 'current_iteration' in data:
            project.current_iteration = data['current_iteration']
        if 'model_metrics' in data:
            project.model_metrics = data['model_metrics']

        db.session.commit()

        return jsonify({
            "message": "Project updated successfully",
            "project": {
                "id": project.id,
                "name": project.name,
                "current_iteration": project.current_iteration,
                "keywords": project.keywords,
                "model_metrics": project.model_metrics
            }
        })

    except Exception as e:
        app.logger.error(f"Error updating project: {str(e)}")
        db.session.rollback()
        return jsonify({"error": "Failed to update project"}), 500


@app.route('/api/admin/projects/<int:project_id>', methods=['DELETE'])
@require_admin()
def delete_project_admin(project_id):
    """Delete project - Admin only"""
    try:
        project = Project.query.get(project_id)
        if not project:
            return jsonify({"error": "Project not found"}), 404

        app.logger.info(f"Admin deleting project {project_id}")

        # Delete all citations first
        citations_deleted = Citation.query.filter_by(project_id=project_id).delete(synchronize_session=False)
        app.logger.info(f"Deleted {citations_deleted} citations for project {project_id}")

        # Delete the project
        db.session.delete(project)
        db.session.commit()

        app.logger.info(f"Successfully deleted project {project_id}")
        return jsonify({"message": "Project deleted successfully"}), 200

    except Exception as e:
        app.logger.error(f"Error deleting project {project_id}: {str(e)}")
        db.session.rollback()
        return jsonify({"error": "Failed to delete project", "details": str(e)}), 500


@app.route('/api/admin/create-admin', methods=['POST'])
def create_admin():
    """Create the first admin user (only works if no admin exists)"""
    try:
        # Check if any admin already exists
        existing_admin = User.query.filter_by(is_admin=True).first()
        if existing_admin:
            return jsonify({"error": "Admin user already exists"}), 400

        data = request.get_json()
        if not all(k in data for k in ["first_name", "last_name", "email", "password"]):
            return jsonify({"error": "Missing required fields"}), 400

        # Check if user with this email already exists
        existing_user = User.query.filter_by(email=data['email']).first()
        if existing_user:
            return jsonify({"error": "Email already registered"}), 400

        admin_user = User(
            first_name=data['first_name'],
            last_name=data['last_name'],
            email=data['email'],
            password=generate_password_hash(data['password']),
            is_admin=True
        )

        db.session.add(admin_user)
        db.session.commit()

        app.logger.info(f"First admin user created: {admin_user.email}")
        return jsonify({"message": "Admin user created successfully"}), 201

    except Exception as e:
        app.logger.error(f"Error creating admin: {str(e)}")
        db.session.rollback()
        return jsonify({"error": "Failed to create admin user"}), 500


@app.route('/api/admin/stats', methods=['GET'])
@require_admin()
def get_admin_stats():
    """Get system statistics - Admin only"""
    try:
        total_users = User.query.count()
        total_admins = User.query.filter_by(is_admin=True).count()
        total_projects = Project.query.count()
        total_citations = Citation.query.count()
        labeled_citations = Citation.query.filter(Citation.is_relevant.isnot(None)).count()

        return jsonify({
            "stats": {
                "total_users": total_users,
                "total_admins": total_admins,
                "total_regular_users": total_users - total_admins,
                "total_projects": total_projects,
                "total_citations": total_citations,
                "labeled_citations": labeled_citations,
                "unlabeled_citations": total_citations - labeled_citations
            }
        })

    except Exception as e:
        app.logger.error(f"Error fetching admin stats: {str(e)}")
        return jsonify({"error": "Failed to fetch statistics"}), 500


def send_reset_email(email, reset_token):
    """Send password reset email using Brevo API"""
    try:
        # Brevo API configuration
        api_key = os.getenv('BREVO_API_KEY')
        sender_email = os.getenv('SENDER_EMAIL', 'noreply@yourapp.com')
        sender_name = os.getenv('SENDER_NAME', 'Literature Review Tool')

        if not api_key:
            app.logger.error("Brevo API key not configured")
            return False

        # Create reset URL - use production URL if available, otherwise development URL
        frontend_url = os.getenv('FRONTEND_URL', 'https://alrt-shilpagopal1.replit.app')

        # If in development mode (when DATABASE_URL contains localhost or is not set for production)
        if not os.getenv('DATABASE_URL') or 'localhost' in os.getenv('DATABASE_URL', ''):
            frontend_url = 'https://2a8d36cb-3602-46e2-82dc-3d70be763e17-00-1bnxhzpxbgva9.janeway.replit.dev'

        reset_url = f"{frontend_url}/reset-password/{reset_token}"

        # Prepare email data
        url = "https://api.brevo.com/v3/smtp/email"

        headers = {
            "accept": "application/json",
            "api-key": api_key,
            "content-type": "application/json"
        }

        payload = {
            "sender": {
                "name": sender_name,
                "email": sender_email
            },
            "to": [
                {
                    "email": email,
                    "name": email.split('@')[0]
                }
            ],
            "subject": "Password Reset Request",
            "htmlContent": f"""
            <html>
            <body>
                <h2>Password Reset Request</h2>
                <p>Hi,</p>
                <p>You requested a password reset for your account.</p>
                <p><a href="{reset_url}" style="background-color: #4CAF50; color: white; padding: 14px 20px; text-decoration: none; border-radius: 4px;">Reset Password</a></p>
                <p>Or copy and paste this link: {reset_url}</p>
                <p>This link will expire in 30 minutes.</p>
                <p>If you didn't request this reset, please ignore this email.</p>
                <br>
                <p>Best regards,<br>Literature Review Tool Team</p>
            </body>
            </html>
            """,
            "textContent": f"""
            Password Reset Request

            Hi,

            You requested a password reset for your account.

            Click the link below to reset your password:
            {reset_url}

            This link will expire in 30 minutes.

            If you didn't request this reset, please ignore this email.

            Best regards,
            Literature Review Tool Team
            """
        }

        # Send email via Brevo API
        response = requests.post(url, json=payload, headers=headers)

        if response.status_code == 201:
            app.logger.info(f"Password reset email sent successfully to {email} via Brevo")
            return True
        else:
            app.logger.error(f"Brevo API error: {response.status_code} - {response.text}")
            return False

    except Exception as e:
        app.logger.error(f"Failed to send email via Brevo: {str(e)}")
        return False


@app.route('/api/auth/forgot-password', methods=['POST'])
def forgot_password():
    try:
        data = request.get_json()

        if not data or 'email' not in data:
            return jsonify({"error": "Email is required"}), 400

        email = data['email'].strip().lower()
        user = User.query.filter_by(email=email).first()

        if not user:
            # For security, don't reveal if email exists or not
            return jsonify({"message": "If an account with this email exists, a password reset link has been sent."}), 200

        # Generate a simple reset token (in production, use JWT or secure tokens)
        import secrets
        reset_token = secrets.token_urlsafe(32)

        # Store the reset token in user record (you'll need to add this field to User model)
        # For now, we'll simulate this with a simple in-memory store
        if not hasattr(app, 'reset_tokens'):
            app.reset_tokens = {}

        # Store token with expiration (30 minutes)
        import time
        app.reset_tokens[reset_token] = {
            'user_id': user.id,
            'email': user.email,
            'expires_at': time.time() + 1800  # 30 minutes
        }

        app.logger.info(f"Password reset token generated for user {user.email}")

        # In a real application, you would send an email here
        if send_reset_email(user.email, reset_token):
            return jsonify({
                "message": "If an account with this email exists, a password reset link has been sent."
            }), 200
        else:
             return jsonify({
                "message": "Failed to send password reset email."
            }), 500

    except Exception as e:
        app.logger.error(f"Forgot password error: {str(e)}")
        return jsonify({"error": "Failed to process request"}), 500


@app.route('/api/auth/reset-password', methods=['POST'])
def reset_password():
    try:
        data = request.get_json()

        if not data or not all(k in data for k in ["token", "new_password"]):
            return jsonify({"error": "Token and new password are required"}), 400

        token = data['token']
        new_password = data['new_password']

        # Validate password strength
        if len(new_password) < 6:
            return jsonify({"error": "Password must be at least 6 characters long"}), 400

        # Check if token exists and is valid
        if not hasattr(app, 'reset_tokens') or token not in app.reset_tokens:
            return jsonify({"error": "Invalid or expired reset token"}), 400

        token_data = app.reset_tokens[token]

        # Check if token has expired
        import time
        if time.time() > token_data['expires_at']:
            del app.reset_tokens[token]
            return jsonify({"error": "Reset token has expired"}), 400

        # Find the user
        user = User.query.get(token_data['user_id'])
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Update the password
        user.password = generate_password_hash(new_password)
        db.session.commit()

        # Remove the used token
        del app.reset_tokens[token]

        app.logger.info(f"Password reset successful for user {user.email}")

        return jsonify({"message": "Password reset successful"}), 200

    except Exception as e:
        app.logger.error(f"Reset password error: {str(e)}")
        db.session.rollback()
        return jsonify({"error": "Failed to reset password"}), 500


@app.route('/api/auth/verify-reset-token', methods=['POST'])
def verify_reset_token():
    try:
        data = request.get_json()

        if not data or 'token' not in data:
            return jsonify({"error": "Token is required"}), 400

        token = data['token']

        # Check if token exists and is valid
        if not hasattr(app, 'reset_tokens') or token not in app.reset_tokens:
            return jsonify({"valid": False, "error": "Invalid reset token"}), 400

        token_data = app.reset_tokens[token]

        # Check if token has expired
        import time
        if time.time() > token_data['expires_at']:
            del app.reset_tokens[token]
            return jsonify({"valid": False, "error": "Reset token has expired"}), 400

        return jsonify({
            "valid": True,
            "email": token_data['email']
        }), 200

    except Exception as e:
        app.logger.error(f"Verify reset token error: {str(e)}")
        return jsonify({"valid": False, "error": "Failed to verify token"}), 500

# for acception of multiple files from users
def validate_file_format(filename):
    """Validate if file format is supported"""
    return filename.lower().endswith(('.csv', '.xlsx', '.xls'))

def process_single_file(file, temp_dir):
    """Process a single file and return DataFrame"""
    try:
        filename = secure_filename(file.filename)
        temp_path = os.path.join(temp_dir, filename)
        file.save(temp_path)
        
        app.logger.info(f"Processing file: {filename}")
        
        if filename.endswith('.csv'):
            df = pd.read_csv(temp_path)
        elif filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(temp_path, engine='openpyxl')
        else:
            raise ValueError(f"Unsupported file format: {filename}")
        
        # Add source file info
        df['source_file'] = filename
        
        # Normalize column names
        df.columns = df.columns.str.lower().str.strip()
        
        # Check required columns
        if 'title' not in df.columns or 'abstract' not in df.columns:
            raise ValueError(f"File {filename} must contain 'title' and 'abstract' columns")
        
        return df, None
        
    except Exception as e:
        return None, str(e)
    finally:
        # Clean up temp file
        if os.path.exists(temp_path):
            os.remove(temp_path)

def merge_multiple_files(files):
    """Process and merge multiple files into a single DataFrame"""
    app.logger.info(f"Processing {len(files)} files for merging")
    
    all_dataframes = []
    processing_errors = []
    file_summaries = []
    
    # Create temporary directory
    temp_dir = '/tmp/multi_upload'
    os.makedirs(temp_dir, exist_ok=True)
    
    try:
        for file in files:
            if not file or file.filename == '':
                continue
                
            if not validate_file_format(file.filename):
                processing_errors.append(f"Invalid format: {file.filename}")
                continue
            
            df, error = process_single_file(file, temp_dir)
            
            if error:
                processing_errors.append(f"Error in {file.filename}: {error}")
                continue
            
            if df is not None and not df.empty:
                all_dataframes.append(df)
                file_summaries.append({
                    'filename': file.filename,
                    'rows': len(df),
                    'columns': list(df.columns)
                })
                app.logger.info(f"Successfully processed {file.filename}: {len(df)} rows")
        
        if not all_dataframes:
            return None, "No valid files processed", []
        
        # Merge all DataFrames
        merged_df = pd.concat(all_dataframes, ignore_index=True, sort=False)
        
        # Standardize columns - ensure title and abstract exist
        required_columns = ['title', 'abstract']
        for col in required_columns:
            if col not in merged_df.columns:
                raise ValueError(f"Missing required column '{col}' after merging")
        
        # Handle optional columns
        optional_mappings = {
            'date': ['date', 'publication_date', 'pub_date', 'year', 'published_year'],
            'authors': ['authors', 'author', 'author_names', 'researchers'],
            'journal': ['journal', 'publication', 'venue', 'source'],
            'doi': ['doi', 'digital_object_identifier'],
            'keywords': ['keywords', 'key_words', 'tags']
        }
        
        # Standardize optional columns
        for standard_name, possible_names in optional_mappings.items():
            found_col = None
            for possible_name in possible_names:
                if possible_name in merged_df.columns:
                    found_col = possible_name
                    break
            
            if found_col and found_col != standard_name:
                merged_df[standard_name] = merged_df[found_col]
        
        app.logger.info(f"Merged {len(all_dataframes)} files into {len(merged_df)} total citations")
        
        return merged_df, None, {
            'files_processed': len(all_dataframes),
            'total_files': len(files),
            'total_citations': len(merged_df),
            'processing_errors': processing_errors,
            'file_summaries': file_summaries
        }
        
    except Exception as e:
        app.logger.error(f"Error merging files: {str(e)}")
        return None, str(e), []
    
    finally:
        # Clean up temp directory
        if os.path.exists(temp_dir):
            import shutil
            shutil.rmtree(temp_dir, ignore_errors=True)

def create_citations_from_merged_data(df, project_id, current_iteration):
    """Create Citation objects from merged DataFrame WITHOUT automatic duplicate processing"""
    app.logger.info(f"Processing {len(df)} citations for project {project_id} - NO automatic duplicate removal")
    
    # Create citations directly without duplicate processing
    new_citations = []
    skipped_count = 0
    
    for _, row in df.iterrows():
        # Validate and clean title
        title_raw = row.get('title')
        if pd.isna(title_raw) or str(title_raw).strip() == '':
            skipped_count += 1
            continue
        
        title_clean = str(title_raw).strip()
        title_clean = re.sub(r'\s+', ' ', title_clean)
        
        # Validate and clean abstract
        abstract_raw = row.get('abstract')
        if pd.isna(abstract_raw) or str(abstract_raw).strip() == '':
            skipped_count += 1
            continue
        
        abstract_clean = str(abstract_raw).strip()
        abstract_clean = re.sub(r'\s+', ' ', abstract_clean)
        
        # Create citation with additional metadata
        citation_data = {
            'title': title_clean,
            'abstract': abstract_clean,
            'project_id': project_id,
            'iteration': current_iteration
        }
        
        # Add optional fields as JSON metadata if they exist
        metadata = {}
        optional_fields = ['date', 'authors', 'journal', 'doi', 'keywords', 'source_file']
        for field in optional_fields:
            if field in row and pd.notna(row[field]):
                metadata[field] = str(row[field]).strip()
        
        citation = Citation(**citation_data)
        new_citations.append(citation)
    
    app.logger.info(f"Citation creation completed: {len(new_citations)} citations created, {skipped_count} skipped")
    return new_citations, 0  # 0 duplicates removed since we're not processing duplicates here

def create_citations_basic(df, project_id, current_iteration):
    """Fallback function for basic citation creation without duplicate processing"""
    new_citations = []
    skipped_count = 0
    
    for _, row in df.iterrows():
        # Validate and clean title
        title_raw = row.get('title')
        if pd.isna(title_raw) or str(title_raw).strip() == '':
            skipped_count += 1
            continue
        
        title_clean = str(title_raw).strip()
        title_clean = re.sub(r'\s+', ' ', title_clean)
        
        # Validate and clean abstract
        abstract_raw = row.get('abstract')
        if pd.isna(abstract_raw) or str(abstract_raw).strip() == '':
            skipped_count += 1
            continue
        
        abstract_clean = str(abstract_raw).strip()
        abstract_clean = re.sub(r'\s+', ' ', abstract_clean)
        
        # Create citation with additional metadata
        citation_data = {
            'title': title_clean,
            'abstract': abstract_clean,
            'project_id': project_id,
            'iteration': current_iteration
        }
        
        # Add optional fields as JSON metadata if they exist
        metadata = {}
        optional_fields = ['date', 'authors', 'journal', 'doi', 'keywords', 'source_file']
        for field in optional_fields:
            if field in row and pd.notna(row[field]):
                metadata[field] = str(row[field]).strip()
        
        citation = Citation(**citation_data)
        
        # Store metadata in a way that doesn't break existing schema
        # You might want to add a metadata JSON column to Citation model
        # For now, we'll just use the core fields
        new_citations.append(citation)
    
    return new_citations, skipped_count



# Update your existing get_projects function
@app.route('/api/projects', methods=['GET'])
def get_projects():
    user_id = request.headers.get('X-User-Id')
    if not user_id:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        user_id_int = int(user_id)

        # Check if user exists and get admin status
        user = User.query.get(user_id_int)
        if not user:
            return jsonify({"error": "User not found"}), 401

        # Admin sees all projects, regular users see only their own
        if user.is_admin:
            projects = Project.query.all()
            app.logger.info(f"Admin user {user_id_int} - Found {len(projects)} total projects")
        else:
            projects = Project.query.filter_by(user_id=user_id_int).all()
            app.logger.info(f"Regular user {user_id_int} - Found {len(projects)} personal projects")

        return jsonify({
            "projects": [{
                "id": p.id,
                "name": p.name,
                "created_at": p.created_at,
                "current_iteration": p.current_iteration,
                "user_id": p.user_id,
                "owner_name": f"{p.user.first_name} {p.user.last_name}" if user.is_admin else None,
                
                # New status fields
                "duplicates_removed": p.duplicates_removed,
                "duplicates_count": p.duplicates_count,
                "keywords_selected": p.keywords_selected,
                "citations_count": p.citations_count
            } for p in projects]
        })
    except Exception as e:
        app.logger.error(f"Error fetching projects for user {user_id}: {str(e)}")
        return jsonify({"error": "Failed to fetch projects"}), 500


@app.route('/api/projects', methods=['POST'])
def create_project():
    try:
        app.logger.info("Received project creation request")
        app.logger.info(f"Headers: {dict(request.headers)}")
        app.logger.info(f"Files: {list(request.files.keys())}")
        app.logger.info(f"Form data: {dict(request.form)}")

        user_id = request.headers.get('X-User-Id')
        if not user_id:
            app.logger.error("No user ID in request headers")
            return jsonify({"error": "Unauthorized"}), 401

        user_id_int = int(user_id)

        # Check if this is a multi-file upload request
        if request.files:
            # Multi-file upload with project creation
            project_name = request.form.get('name')
            if not project_name:
                return jsonify({"error": "Project name is required"}), 400

            # Get all uploaded files
            uploaded_files = []
            for key in request.files:
                files = request.files.getlist(key)
                uploaded_files.extend(files)

            if not uploaded_files:
                return jsonify({"error": "At least one file is required"}), 400

            app.logger.info(f"Processing {len(uploaded_files)} files for project: {project_name}")

            # Process and merge multiple files
            merged_df, merge_error, merge_summary = merge_multiple_files(uploaded_files)

            if merge_error:
                return jsonify({
                    "error": "Failed to process files",
                    "details": merge_error,
                    "summary": merge_summary
                }), 400

            # Create the project
            project = Project(
                name=project_name,
                user_id=user_id_int,
                keywords={"include": [], "exclude": []}
            )
            db.session.add(project)
            db.session.commit()
            db.session.refresh(project)

            # Create citations from merged data
            new_citations, skipped_count = create_citations_from_merged_data(
                merged_df, project.id, project.current_iteration
            )

            # Bulk insert citations
            if new_citations:
                db.session.bulk_save_objects(new_citations)
                
                # Update project status
                project.citations_count = len(new_citations)
                project.duplicates_removed = False
                project.keywords_selected = False
                
                db.session.commit()

            app.logger.info(f"Project created successfully with ID: {project.id}, {len(new_citations)} citations added")

            return jsonify({
                "project": {
                    "id": project.id,
                    "name": project.name,
                    "created_at": project.created_at,
                    "current_iteration": project.current_iteration,
                    "citations_count": len(new_citations)
                },
                "upload_summary": {
                    "files_processed": merge_summary.get('files_processed', 0),
                    "total_files": merge_summary.get('total_files', 0),
                    "total_citations": len(new_citations),
                    "skipped_citations": skipped_count,
                    "processing_errors": merge_summary.get('processing_errors', []),
                    "file_summaries": merge_summary.get('file_summaries', [])
                },
                "message": f"Project created successfully with {len(new_citations)} citations from {merge_summary.get('files_processed', 0)} files"
            }), 201

        else:
            # Regular project creation without files (existing functionality)
            data = request.get_json()
            if not data:
                return jsonify({"error": "No data provided"}), 400

            if 'name' not in data:
                return jsonify({"error": "Project name is required"}), 400

            project = Project(
                name=data['name'],
                user_id=user_id_int,
                keywords={"include": [], "exclude": []}
            )
            db.session.add(project)
            db.session.commit()
            db.session.refresh(project)

            return jsonify({
                "project": {
                    "id": project.id,
                    "name": project.name,
                    "created_at": project.created_at,
                    "current_iteration": project.current_iteration
                }
            }), 201

    except Exception as e:
        app.logger.error(f"Project creation error: {str(e)}")
        db.session.rollback()
        return jsonify({"error": "Failed to create project", "details": str(e)}), 500

@app.route('/api/projects/<int:project_id>/citations', methods=['POST'])
def add_citations(project_id):
    try:
        app.logger.info(f"Received citation upload request for project {project_id}")
        app.logger.info(f"Request headers: {dict(request.headers)}")
        app.logger.info(f"Request files: {request.files}")
        app.logger.info(f"Request form: {request.form}")

        user_id = request.headers.get('X-User-Id')
        if not user_id:
            app.logger.error("No user ID in request headers")
            return jsonify({"error": "Unauthorized"}), 401

        user_id_int = int(user_id)
        user = User.query.get(user_id_int)
        if not user:
            return jsonify({"error": "User not found"}), 401

        # Allow admin access to any project, regular users only their own
        if user.is_admin:
            project = Project.query.get(project_id)
        else:
            project = Project.query.filter_by(id=project_id, user_id=user_id_int).first()

        if not project:
            app.logger.error(f"Project {project_id} not found for user {user_id}")
            return jsonify({"error": "Project not found"}), 404

        if 'file' not in request.files:
            app.logger.info("No file in request, checking for JSON data")
            data = request.get_json()
            if not isinstance(data.get('citations'), list):
                return jsonify({"error": "Citations must be a list"}), 400

            new_citations = []
            for citation_data in data['citations']:
                citation = Citation(title=citation_data['title'],
                                    abstract=citation_data['abstract'],
                                    project_id=project_id,
                                    iteration=project.current_iteration)
                new_citations.append(citation)
        else:
            file = request.files['file']
            app.logger.info(f"Received file: {file.filename if file else 'No file'}")

            if not file or file.filename == '':
                app.logger.error("No file provided or empty filename")
                return jsonify({"error": "No file provided"}), 400

            if not file.filename.endswith(('.csv', '.xlsx')):
                app.logger.error(f"Invalid file format: {file.filename}")
                return jsonify({
                    "error": "Invalid file format. Only CSV and Excel files are supported.",
                    "filename": file.filename
                }), 400

            app.logger.info(f"Processing file: {file.filename} ({file.content_length} bytes)")

            try:
                app.logger.info(f"Attempting to process file: {file.filename}")
                # Save the file temporarily
                temp_path = os.path.join('/tmp', secure_filename(file.filename))
                file.save(temp_path)
                app.logger.info(f"Saved file temporarily: {temp_path}")

                try:
                    if file.filename.endswith('.csv'):
                        app.logger.info("Processing CSV file")
                        df = pd.read_csv(temp_path)
                    elif file.filename.endswith('.xlsx'):
                        app.logger.info("Processing Excel file")
                        df = pd.read_excel(temp_path, engine='openpyxl')
                        app.logger.info(f"Excel file read successfully with {len(df)} rows")
                    else:
                        app.logger.error(f"Unsupported file format: {file.filename}")
                        return jsonify({
                            "error": "Unsupported file format",
                            "details": f"File {file.filename} is not supported. Only .csv and .xlsx files are allowed"
                        }), 400
                    # Check file size limits for performance
                    if len(df) > 50000:
                        return jsonify({
                            "error": "File too large",
                            "details": f"File contains {len(df)} rows. Maximum supported is 50,000 rows for optimal performance."
                        }), 400

                    # Clean up temp file
                    os.remove(temp_path)
                except Exception as e:
                    app.logger.error(f"File read error: {str(e)}")
                    return jsonify({
                        "error": "Failed to read file",
                        "details": str(e)
                    }), 400

                if df.empty:
                    return jsonify({"error": "File contains no data"}), 400

                # Normalize column names to lowercase for case-insensitive validation
                original_columns = list(df.columns)
                df.columns = df.columns.str.lower().str.strip()

                # Check for required columns (case-insensitive)
                if 'title' not in df.columns or 'abstract' not in df.columns:
                    return jsonify({
                        "error": "File must contain 'title' and 'abstract' columns (case-insensitive)",
                        "found_columns": original_columns,
                        "normalized_columns": list(df.columns)
                    }), 400

                # Create citations directly without duplicate processing
                app.logger.info(f"Processing {len(df)} citations WITHOUT automatic duplicate detection")
                
                new_citations = []
                all_metadata_keys = set()
                for _, row in df.iterrows():
                    # Normalize title: strip spaces, clean special characters, handle NaN
                    title_raw = row['title']
                    if pd.isna(title_raw) or str(title_raw).strip() == '':
                        continue  # Skip citations with empty titles

                    title_clean = str(title_raw).strip()
                    # Remove excessive whitespace and normalize
                    title_clean = re.sub(r'\s+', ' ', title_clean)

                    # Normalize abstract: strip spaces, clean special characters, handle NaN
                    abstract_raw = row['abstract']
                    if pd.isna(abstract_raw):
                        abstract_clean = ""
                    else:
                        abstract_clean = str(abstract_raw).strip()
                        # Remove excessive whitespace and normalize
                        abstract_clean = re.sub(r'\s+', ' ', abstract_clean)

                    # Skip citations with empty abstracts as they're not useful
                    if abstract_clean == "":
                        continue

                    # Build metadata from remaining columns
                    row_dict = {k: (None if pd.isna(v) else v) for k, v in row.items()}
                    core_fields = {'title', 'abstract'}
                    metadata = {k: str(v) for k, v in row_dict.items() if k not in core_fields and v is not None}
                    all_metadata_keys.update(metadata.keys())

                    citation = Citation(
                        title=title_clean,
                        abstract=abstract_clean,
                        project_id=project_id,
                        iteration=project.current_iteration,
                        extra_metadata=metadata
                    )
                    new_citations.append(citation)

                # Update project status after successful upload
                project.citations_count = len(new_citations)
                # Reset status flags when new citations are uploaded
                project.duplicates_removed = False
                project.keywords_selected = bool(project.keywords.get('include') or project.keywords.get('exclude'))
                
                db.session.bulk_save_objects(new_citations)
                # Update available_columns cache on project
                available = sorted(set(list(all_metadata_keys) + ['title', 'abstract']))
                project.available_columns = available
                db.session.commit()

                # Prepare response message for citations added without duplicate processing
                response_message = f"Added {len(new_citations)} citations (duplicates will be handled manually)"
                
                return jsonify({
                    "message": response_message,
                    "total_citations": len(new_citations),
                    "note": "Use the 'Remove Duplicates' button to process duplicates manually",
                    
                    # Status information for frontend
                    "citations_count": len(new_citations),
                    "duplicates_removed": False,  # No automatic duplicate removal
                    "keywords_selected": project.keywords_selected,
                    
                    # Enhanced duplicate processing details
                    "duplicate_processing": {
                        "success": duplicates_removed_status,
                        "duplicates_removed": duplicate_result.get('duplicates_removed', 0) if not duplicate_result.get('error') else 0,
                        "strategy": duplicate_result.get('strategy', 'None') if not duplicate_result.get('error') else 'Fallback',
                        "columns_standardized": duplicate_result.get('error') is None
                    }
                }), 201

            except Exception as e:
                db.session.rollback()
                app.logger.error(f"Upload processing error: {str(e)}")
                return jsonify({
                    "error": "Failed to process upload",
                    "details": str(e)
                }), 500

    except Exception as e:
        app.logger.error(f"General error in add_citations: {str(e)}")
        return jsonify({
            "error": "Internal server error",
            "details": str(e)
        }), 500


@app.route('/api/projects/<int:project_id>/citations/<int:citation_id>/relevance',
           methods=['PUT'])
def update_citation(project_id, citation_id):
    try:
        user_id = request.headers.get('X-User-Id')
        if not user_id:
            return jsonify({"error": "Unauthorized"}), 401

        user_id_int = int(user_id)
        user = User.query.get(user_id_int)
        if not user:
            return jsonify({"error": "User not found"}), 401

        # Allow admin access to any project, regular users only their own
        if user.is_admin:
            project = Project.query.get(project_id)
        else:
            project = Project.query.filter_by(id=project_id, user_id=user_id_int).first()

        if not project:
            return jsonify({"error": "Project not found"}), 404

        citation = Citation.query.filter_by(id=citation_id,
                                          project_id=project_id).first()
        if not citation:
            return jsonify({"error": "Citation not found"}), 404

        data = request.get_json()
        if 'is_relevant' in data:
            # Allow null/None value to unlabel the citation
            citation.is_relevant = data['is_relevant'] if data['is_relevant'] in [True, False] else None
            if citation.is_relevant is not None:
                citation.iteration = project.current_iteration
            db.session.commit()
            status = "unlabeled" if citation.is_relevant is None else str(citation.is_relevant)
            app.logger.info(f"Updated citation {citation_id} relevance to {status} for iteration {citation.iteration}")

        return jsonify({
            "citation": {
                "id": citation.id,
                "title": citation.title,
                "abstract": citation.abstract,
                "is_relevant": citation.is_relevant,
                "iteration": citation.iteration
            }
        })
    except Exception as e:
        app.logger.error(f"Error updating citation: {str(e)}")
        db.session.rollback()
        return jsonify({"error": "Failed to update citation"}), 500


@app.route('/api/projects/<int:project_id>/citations/<int:citation_id>/duplicate-status',
           methods=['PUT'])
def toggle_citation_duplicate_status(project_id, citation_id):
    """Toggle citation duplicate status between true and false"""
    try:
        user_id = request.headers.get('X-User-Id')
        if not user_id:
            return jsonify({"error": "Unauthorized"}), 401

        user_id_int = int(user_id)
        user = User.query.get(user_id_int)
        if not user:
            return jsonify({"error": "User not found"}), 401

        # Allow admin access to any project, regular users only their own
        if user.is_admin:
            project = Project.query.get(project_id)
        else:
            project = Project.query.filter_by(id=project_id, user_id=user_id_int).first()

        if not project:
            return jsonify({"error": "Project not found"}), 404

        citation = Citation.query.filter_by(id=citation_id,
                                          project_id=project_id).first()
        if not citation:
            return jsonify({"error": "Citation not found"}), 404

        # Toggle the duplicate status
        current_status = getattr(citation, 'is_duplicate', False)
        new_status = not current_status
        citation.is_duplicate = new_status

        # Update project counts
        if new_status:
            # Citation is now marked as duplicate
            project.duplicates_count = getattr(project, 'duplicates_count', 0) + 1
            project.citations_count = getattr(project, 'citations_count', 0) - 1
        else:
            # Citation is now restored (not duplicate)
            project.duplicates_count = max(0, getattr(project, 'duplicates_count', 0) - 1)
            project.citations_count = getattr(project, 'citations_count', 0) + 1

        db.session.commit()
        
        status_text = "marked as duplicate" if new_status else "restored (not duplicate)"
        app.logger.info(f"Updated citation {citation_id} duplicate status to {new_status} for project {project_id}")

        return jsonify({
            "message": f"Citation {status_text} successfully",
            "citation": {
                "id": citation.id,
                "title": citation.title,
                "abstract": citation.abstract,
                "is_duplicate": citation.is_duplicate,
                "iteration": citation.iteration
            },
            "project": {
                "citations_count": project.citations_count,
                "duplicates_count": project.duplicates_count
            }
        })
        
    except Exception as e:
        app.logger.error(f"Error toggling citation duplicate status: {str(e)}")
        db.session.rollback()
        return jsonify({"error": "Failed to toggle citation duplicate status"}), 500


@app.route('/api/projects/<int:project_id>/train', methods=['POST'])
def train_model(project_id):
    try:
        user_id = request.headers.get('X-User-Id')
        if not user_id:
            return jsonify({"error": "Unauthorized"}), 401

        user_id_int = int(user_id)
        user = User.query.get(user_id_int)
        if not user:
            return jsonify({"error": "User not found"}), 401

        # Allow admin access to any project, regular users only their own
        if user.is_admin:
            project = Project.query.get(project_id)
        else:
            project = Project.query.filter_by(id=project_id, user_id=user_id_int).first()

        if not project:
            return jsonify({"error": "Project not found"}), 404

        # Get only user-labeled citations (exclude duplicates and keyword-excluded citations)
        user_labeled_citations = Citation.query.filter(
            Citation.project_id == project_id,
            Citation.is_relevant.isnot(None),
            Citation.is_duplicate == False  # Exclude duplicates from training
        ).all()

        # Filter out citations that would be excluded by keywords but keep user labels intact
        filtered_citations = []
        keyword_excluded_count = 0

        for citation in user_labeled_citations:
            should_exclude = False
            text = f"{citation.title} {citation.abstract}".lower()

            for exclude_kw in project.keywords.get('exclude', []):
                word = exclude_kw['word'].lower()
                frequency = exclude_kw.get('frequency', 1)
                occurrences = text.count(word)

                if occurrences >= frequency:
                    should_exclude = True
                    keyword_excluded_count += 1
                    break

            if not should_exclude:
                filtered_citations.append(citation)

        if len(filtered_citations) < 10:
            return jsonify({
                "error": f"Need at least 10 user-labeled citations (excluding duplicates and keyword-filtered) to train. Currently have {len(filtered_citations)} valid labeled citations. Total user-labeled: {len(user_labeled_citations)}, Keyword-excluded: {keyword_excluded_count}"
            }), 400

        # Count relevant and irrelevant citations from filtered set
        relevant_count = sum(1 for c in filtered_citations if c.is_relevant == True)
        irrelevant_count = sum(1 for c in filtered_citations if c.is_relevant == False)

        # NEW validation:
        MIN_LABELS_PER_CATEGORY = 5
        MAX_LABELS_PER_CATEGORY = 10

        if relevant_count < MIN_LABELS_PER_CATEGORY or irrelevant_count < MIN_LABELS_PER_CATEGORY:
            return jsonify({
                "error": f"Need at least {MIN_LABELS_PER_CATEGORY} relevant and {MIN_LABELS_PER_CATEGORY} irrelevant user-labeled citations (excluding duplicates and keyword-filtered). Currently have {relevant_count} relevant and {irrelevant_count} irrelevant from {len(filtered_citations)} valid citations."
            }), 400

        if relevant_count > MAX_LABELS_PER_CATEGORY or irrelevant_count > MAX_LABELS_PER_CATEGORY:
            return jsonify({
                "error": f"Maximum {MAX_LABELS_PER_CATEGORY} citations per category allowed for balanced training. Currently have {relevant_count} relevant and {irrelevant_count} irrelevant. Please unlabel some citations before training."
            }), 400

        # Update the minimum total citations check:
        if len(filtered_citations) < (MIN_LABELS_PER_CATEGORY * 2):
            return jsonify({
                "error": f"Need at least {MIN_LABELS_PER_CATEGORY * 2} user-labeled citations total (excluding duplicates and keyword-filtered) to train. Currently have {len(filtered_citations)} valid labeled citations."
            }), 400

        app.logger.info(f"Training model for project {project_id} with {len(filtered_citations)} user-labeled citations ({relevant_count} relevant, {irrelevant_count} irrelevant). Excluded {keyword_excluded_count} due to keywords.")

        review_system = LiteratureReviewSystem(project_id)
        result = review_system.train_iteration(project.current_iteration)

        if 'error' in result:
            app.logger.error(f"Training error for project {project_id}: {result['error']}")
            return jsonify(result), 400

        # Add filtering metadata to result
        result['filtering_info'] = {
            'total_user_labeled': len(user_labeled_citations),
            'keyword_excluded': keyword_excluded_count,
            'duplicates_excluded': len(user_labeled_citations) - len(filtered_citations) - keyword_excluded_count,
            'used_for_training': len(filtered_citations)
        }

        # Store metrics and increment iteration
        if str(project.current_iteration) not in project.model_metrics:
            project.model_metrics[str(project.current_iteration)] = {}
        project.model_metrics[str(project.current_iteration)] = result['metrics']
        project.current_iteration += 1
        db.session.commit()

        app.logger.info(f"Successfully trained model for project {project_id}, moving to iteration {project.current_iteration}")
        return jsonify(result)

    except Exception as e:
        app.logger.error(f"Training model error for project {project_id}: {str(e)}")
        db.session.rollback()
        return jsonify({
            "error": "Failed to train model",
            "details": str(e)
        }), 500


@app.route('/api/projects/<int:project_id>/keywords', methods=['GET'])
def get_keywords(project_id):
    user_id = request.headers.get('X-User-Id')
    if not user_id:
        return jsonify({"error": "Unauthorized"}), 401

    user_id_int = int(user_id)
    user = User.query.get(user_id_int)
    if not user:
        return jsonify({"error": "User not found"}), 401

    # Allow admin access to any project, regular users only their own
    if user.is_admin:
        project = Project.query.get(project_id)
    else:
        project = Project.query.filter_by(id=project_id, user_id=user_id_int).first()

    if not project:
        return jsonify({"error": "Project not found"}), 404

    # Exclude citations marked as duplicates
    citations = Citation.query.filter_by(project_id=project_id, is_duplicate=False).all()
    if not citations:
        return jsonify({"error": "No citations found"}), 404

    # Combine titles and abstracts for TF-IDF
    texts = [f"{c.title} {c.abstract}" for c in citations]
    vectorizer = TfidfVectorizer(max_features=70, stop_words='english')
    tfidf_matrix = vectorizer.fit_transform(texts)

    # Get feature names and their scores
    feature_names = vectorizer.get_feature_names_out()
    scores = tfidf_matrix.sum(axis=0).A1

    # Get top 50 keywords by TF-IDF score but only return the words
    top_indices = scores.argsort()[-50:][::-1]  
    suggested_keywords = [{"word": feature_names[i]} for i in top_indices]

    return jsonify({
        "suggested_keywords": suggested_keywords,
        "selected_keywords": project.keywords
    })


@app.route('/api/projects/<int:project_id>/keywords', methods=['PUT'])
def update_keywords(project_id):
    user_id = request.headers.get('X-User-Id')
    if not user_id:
        return jsonify({"error": "Unauthorized"}), 401

    user_id_int = int(user_id)
    user = User.query.get(user_id_int)
    if not user:
        return jsonify({"error": "User not found"}), 401

    # Allow admin access to any project, regular users only their own
    if user.is_admin:
        project = Project.query.get(project_id)
    else:
        project = Project.query.filter_by(id=project_id, user_id=user_id_int).first()

    if not project:
        return jsonify({"error": "Project not found"}), 404

    data = request.get_json()
    if not isinstance(data, dict) or not all(k in data for k in ['include', 'exclude']):
        return jsonify({"error": "Invalid keywords format"}), 400

    # Validate exclude keywords format
    if not all(isinstance(item, dict) and 'word' in item and 'frequency' in item 
              for item in data['exclude']):
        return jsonify({"error": "Invalid exclude keywords format"}), 400

    # Set default frequency of 1 if not specified
    for item in data['exclude']:
        if not isinstance(item['frequency'], int) or item['frequency'] < 1:
            item['frequency'] = 1

    project.keywords = data
    
    # Update keywords_selected status
    has_keywords = bool(data.get('include') or data.get('exclude'))
    project.keywords_selected = has_keywords
    
    db.session.commit()

    return jsonify({
        "keywords": project.keywords,
        "keywords_selected": project.keywords_selected
    })


@app.route('/api/projects/<int:project_id>/duplicate-options', methods=['GET', 'OPTIONS'])
def get_duplicate_options(project_id):
    # Handle preflight
    if request.method == 'OPTIONS':
        return ('', 200)

    user_id = request.headers.get('X-User-Id')
    if not user_id:
        return jsonify({"error": "Unauthorized"}), 401

    user_id_int = int(user_id)
    user = User.query.get(user_id_int)
    if not user:
        return jsonify({"error": "User not found"}), 401

    # Allow admin access to any project, regular users only their own
    if user.is_admin:
        project = Project.query.get(project_id)
    else:
        project = Project.query.filter_by(id=project_id, user_id=user_id_int).first()

    if not project:
        return jsonify({"error": "Project not found"}), 404

    # Fallbacks
    available_columns = project.available_columns or ['title', 'abstract']
    current_config = project.duplicate_config or {"mode": "default", "columns": [], "threshold": 0.9}

    return jsonify({
        "available_columns": available_columns,
        "current_config": current_config
    })


@app.route('/api/projects/<int:project_id>/citations/filter', methods=['GET'])
def filter_citations(project_id):
    user_id = request.headers.get('X-User-Id')
    if not user_id:
        return jsonify({"error": "Unauthorized"}), 401

    user_id_int = int(user_id)
    user = User.query.get(user_id_int)
    if not user:
        return jsonify({"error": "User not found"}), 401

    # Allow admin access to any project, regular users only their own
    if user.is_admin:
        project = Project.query.get(project_id)
    else:
        project = Project.query.filter_by(id=project_id, user_id=user_id_int).first()

    if not project:
        return jsonify({"error": "Project not found"}), 404

    # Get query parameters
    iteration = request.args.get('iteration', type=int)
    is_relevant = request.args.get('is_relevant', type=lambda v: v.lower() == 'true' if v else None)
    sort_order = request.args.get('sort_order', 'desc')  # NEW: desc (default) or asc
    
    # Validate sort_order parameter
    if sort_order not in ['asc', 'desc']:
        sort_order = 'desc'

    query = Citation.query.filter_by(project_id=project_id, is_duplicate=False)
    if iteration is not None:
        query = query.filter_by(iteration=iteration)
    if is_relevant is not None:
        query = query.filter_by(is_relevant=is_relevant)

    citations = query.all()

    # Apply keyword filtering
    filtered_citations = apply_keyword_filtering(citations, project)

    # Build included/excluded sets
    included_ids = {c.id for c in filtered_citations}

    # NEW: Get relevance scores and sort accordingly
    if filtered_citations:
        try:
            review_system = LiteratureReviewSystem(project_id)
            predictions = review_system.predict_relevance([{
                'title': c.title,
                'abstract': c.abstract
            } for c in filtered_citations])

            # Combine citations with their relevance scores
            citations_with_scores = []
            for citation, prediction in zip(filtered_citations, predictions):
                relevance_score = prediction.get('relevance_probability', 0) if 'error' not in prediction else 0
                citations_with_scores.append({
                    'citation': citation,
                    'relevance_score': relevance_score
                })

            # Sort based on sort_order parameter
            reverse_sort = (sort_order == 'desc')
            citations_with_scores.sort(key=lambda x: x['relevance_score'], reverse=reverse_sort)
            
            # Extract sorted citations
            sorted_citations = [item['citation'] for item in citations_with_scores]
            
            # Prepare included (sorted) with scores
            included_response = []
            for item in citations_with_scores:
                citation = item['citation']
                included_response.append({
                    "id": citation.id,
                    "title": citation.title,
                    "abstract": citation.abstract,
                    "is_relevant": citation.is_relevant,
                    "iteration": citation.iteration,
                    "is_duplicate": getattr(citation, 'is_duplicate', False),
                    "relevance_score": round(item['relevance_score'], 4),
                    "included_by_keywords": True
                })

            # Append excluded (no scores)
            excluded_response = [{
                "id": c.id,
                "title": c.title,
                "abstract": c.abstract,
                "is_relevant": c.is_relevant,
                "iteration": c.iteration,
                "is_duplicate": getattr(c, 'is_duplicate', False),
                "relevance_score": None,
                "included_by_keywords": False
            } for c in citations if c.id not in included_ids]

            citations_response = included_response + excluded_response

        except Exception as e:
            app.logger.warning(f"Could not calculate relevance scores: {str(e)}")
            # Fallback: all citations, included flagged, no scores
            citations_response = [{
                "id": c.id,
                "title": c.title,
                "abstract": c.abstract,
                "is_relevant": c.is_relevant,
                "iteration": c.iteration,
                "is_duplicate": getattr(c, 'is_duplicate', False),
                "relevance_score": None,
                "included_by_keywords": (c.id in included_ids)
            } for c in citations]
    else:
        # No citations after filters: return empty
        citations_response = []

    return jsonify({
        "citations": citations_response,  # all non-duplicate citations with included_by_keywords flag
        "total_before_filtering": len(citations),
        "total_after_filtering": len(filtered_citations),
        "filtered_by_keywords": len(citations) - len(filtered_citations),
        "sort_order": sort_order,  # NEW: Include current sort order in response
        "sort_available": any(item.get('relevance_score') is not None for item in citations_response)
    })



@app.route('/api/projects/<int:project_id>/download', methods=['GET'])
def download_results(project_id):
    user_id = request.headers.get('X-User-Id')
    if not user_id:
        return jsonify({"error": "Unauthorized"}), 401

    user_id_int = int(user_id)
    user = User.query.get(user_id_int)
    if not user:
        return jsonify({"error": "User not found"}), 401

    # Allow admin access to any project, regular users only their own
    if user.is_admin:
        project = Project.query.get(project_id)
    else:
        project = Project.query.filter_by(id=project_id, user_id=user_id_int).first()

    if not project:
        return jsonify({"error": "Project not found"}), 404

    # NEW: Get sort_order parameter
    sort_order = request.args.get('sort_order', 'desc')
    if sort_order not in ['asc', 'desc']:
        sort_order = 'desc'

    citations = Citation.query.filter_by(project_id=project_id).all()

    output = io.BytesIO()
    from openpyxl import Workbook
    workbook = Workbook()
    worksheet = workbook.active

    # Write headers
    headers = [
        'Title', 'Abstract', 'Is Relevant', 'Iteration', 'Relevance Score'
    ]
    for col, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col, value=header)

    # Get relevance scores for citations
    review_system = LiteratureReviewSystem(project_id)
    predictions = review_system.predict_relevance([{
        'title': c.title,
        'abstract': c.abstract
    } for c in citations])

    # Create list of data and sort by relevance score
    data_rows = []
    for citation, prediction in zip(citations, predictions):
        relevance_score = prediction.get('relevance_probability', 0) if 'error' not in prediction else 0
        data_rows.append({
            'title': citation.title,
            'abstract': citation.abstract,
            'is_relevant': 'Yes' if citation.is_relevant else 'No' if citation.is_relevant is not None else 'Unclassified',
            'iteration': citation.iteration,
            'relevance_score': relevance_score
        })

    # NEW: Sort by relevance score based on sort_order parameter
    reverse_sort = (sort_order == 'desc')
    data_rows.sort(key=lambda x: x['relevance_score'], reverse=reverse_sort)

    # Write sorted data
    for row, data in enumerate(data_rows, start=2):
        worksheet.cell(row=row, column=1, value=data['title'])
        worksheet.cell(row=row, column=2, value=data['abstract'])
        worksheet.cell(row=row, column=3, value=data['is_relevant'])
        worksheet.cell(row=row, column=4, value=data['iteration'])
        worksheet.cell(row=row, column=5, value=data['relevance_score'])

    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'project_{project_id}_results_{sort_order}.xlsx')

#  Add a new endpoint to get current sort capabilities
@app.route('/api/projects/<int:project_id>/sort-info', methods=['GET'])
def get_sort_info(project_id):
    """Get information about sorting capabilities for this project"""
    user_id = request.headers.get('X-User-Id')
    if not user_id:
        return jsonify({"error": "Unauthorized"}), 401

    user_id_int = int(user_id)
    user = User.query.get(user_id_int)
    if not user:
        return jsonify({"error": "User not found"}), 401

    # Allow admin access to any project, regular users only their own
    if user.is_admin:
        project = Project.query.get(project_id)
    else:
        project = Project.query.filter_by(id=project_id, user_id=user_id_int).first()

    if not project:
        return jsonify({"error": "Project not found"}), 404

    # Check if model is trained (has metrics for any iteration)
    has_trained_model = bool(project.model_metrics)
    
    return jsonify({
        "sort_available": has_trained_model,
        "current_iteration": project.current_iteration,
        "model_trained": has_trained_model,
        "sort_options": [
            {"value": "desc", "label": "High to Low Relevance (Default)", "description": "Best for finding obviously irrelevant papers"},
            {"value": "asc", "label": "Low to High Relevance", "description": "Best for finding unexpectedly relevant papers"}
        ]
    })


@app.route('/api/projects/<int:project_id>/predict', methods=['POST'])
def predict_citations(project_id):
    user_id = request.headers.get('X-User-Id')
    if not user_id:
        return jsonify({"error": "Unauthorized"}), 401

    user_id_int = int(user_id)
    user = User.query.get(user_id_int)
    if not user:
        return jsonify({"error": "User not found"}), 401

    # Allow admin access to any project, regular users only their own
    if user.is_admin:
        project = Project.query.get(project_id)
    else:
        project = Project.query.filter_by(id=project_id, user_id=user_id_int).first()

    if not project:
        return jsonify({"error": "Project not found"}), 404

    data = request.get_json()
    if not isinstance(data.get('citations'), list):
        return jsonify({"error": "Citations must be a list"}), 400

    review_system = LiteratureReviewSystem(project_id)
    predictions = review_system.predict_relevance(data['citations'])

    return jsonify({"predictions": predictions})


@app.route('/api/projects/<int:project_id>/iterations', methods=['GET'])
def get_iteration_info(project_id):
    user_id = request.headers.get('X-User-Id')
    if not user_id:
        return jsonify({"error": "Unauthorized"}), 401

    user_id_int = int(user_id)
    user = User.query.get(user_id_int)
    if not user:
        return jsonify({"error": "User not found"}), 401

    # Allow admin access to any project, regular users only their own
    if user.is_admin:
        project = Project.query.get(project_id)
    else:
        project = Project.query.filter_by(id=project_id, user_id=user_id_int).first()

    if not project:
        return jsonify({"error": "Project not found"}), 404

    return jsonify({
        "current_iteration": project.current_iteration,
        "max_iterations": 10,
        "metrics": project.model_metrics
    })


def add_duplicate_status_to_details(duplicate_details, project_id):
    """Add is_duplicate field to duplicate details based on current database state"""
    try:
        # Get all citations for this project to check their current duplicate status
        citations = Citation.query.filter_by(project_id=project_id).all()
        citation_status_map = {c.id: getattr(c, 'is_duplicate', False) for c in citations}
        
        # Add is_duplicate field to each duplicate detail
        enhanced_details = []
        for detail in duplicate_details:
            enhanced_detail = detail.copy()
            
            # Add is_duplicate status to kept citation
            if 'kept' in detail and 'id' in detail['kept']:
                kept_id = detail['kept']['id']
                enhanced_detail['kept']['is_duplicate'] = citation_status_map.get(kept_id, False)
            
            # Add is_duplicate status to removed citation
            if 'removed' in detail and 'id' in detail['removed']:
                removed_id = detail['removed']['id']
                enhanced_detail['removed']['is_duplicate'] = citation_status_map.get(removed_id, False)
            
            enhanced_details.append(enhanced_detail)
        
        return enhanced_details
        
    except Exception as e:
        app.logger.error(f"Error adding duplicate status to details: {str(e)}")
        # Return original details if enhancement fails
        return duplicate_details


@app.route('/api/projects/<int:project_id>', methods=['GET'])
def get_project_details(project_id):
    """Get project details including citations"""
    try:
        user_id = request.headers.get('X-User-Id')
        if not user_id:
            return jsonify({"error": "Unauthorized"}), 401

        user_id_int = int(user_id)

        # Check if user is admin
        user = User.query.get(user_id_int)
        if not user:
            return jsonify({"error": "User not found"}), 401

        # Allow admin access to any project, regular users only their own
        if user.is_admin:
            project = Project.query.get(project_id)
        else:
            project = Project.query.filter_by(id=project_id, user_id=user_id_int).first()

        if not project:
            return jsonify({"error": "Project not found"}), 404

        citations = Citation.query.filter_by(project_id=project_id).all()
        labeled_count = Citation.query.filter(
            Citation.project_id == project_id,
            Citation.is_relevant.isnot(None)
        ).count()

        # Apply keyword filtering to citations
        non_duplicate_citations = [c for c in citations if not getattr(c, 'is_duplicate', False)]
        filtered_citations = apply_keyword_filtering(non_duplicate_citations, project)

        # Count labeled citations from filtered set
        filtered_labeled_count = sum(1 for c in filtered_citations if c.is_relevant is not None)

        return jsonify({
            "project": {
                "id": project.id,
                "name": project.name,
                "created_at": project.created_at,
                "current_iteration": project.current_iteration,
                "keywords": project.keywords,
                "model_metrics": project.model_metrics,
                "citations_count": len(filtered_citations),
                "labeled_count": filtered_labeled_count,
                "total_uploaded": len(citations),
                "duplicates_count": sum(1 for c in citations if getattr(c, 'is_duplicate', False)),
                "keyword_filtered_count": len(non_duplicate_citations) - len(filtered_citations),
                
                # Existing status fields
                "duplicates_removed": project.duplicates_removed,
                "keywords_selected": project.keywords_selected,
                "citations_count_stored": project.citations_count,
                "duplicates_count_stored": project.duplicates_count,
                
                # NEW: Duplicate details fields with is_duplicate status
                "duplicate_details": add_duplicate_status_to_details(project.duplicate_details or [], project_id) if project.duplicate_details else [],
                "processing_summary": project.processing_summary or {},
                "removal_strategy": project.removal_strategy
            },
            "citations": [{
                "id": c.id,
                "title": c.title,
                "abstract": c.abstract,
                "is_relevant": c.is_relevant,
                "iteration": c.iteration,
                "is_duplicate": getattr(c, 'is_duplicate', False)
            } for c in filtered_citations]
        })

    except Exception as e:
        app.logger.error(f"Error fetching project details: {str(e)}")
        return jsonify({"error": "Failed to fetch project details"}), 500


@app.route('/api/projects/<int:project_id>', methods=['DELETE'])
def delete_project(project_id):
    try:
        user_id = request.headers.get('X-User-Id')
        if not user_id:
            app.logger.error("No X-User-Id header provided")
            return jsonify({"error": "Unauthorized"}), 401

        try:
            user_id_int = int(user_id)
        except ValueError:
            app.logger.error(f"Invalid user ID format: {user_id}")
            return jsonify({"error": "Invalid user ID"}), 401

        # Check if user exists and get admin status
        user = User.query.get(user_id_int)
        if not user:
            app.logger.error(f"User {user_id_int} not found")
            return jsonify({"error": "User not found"}), 401

        # Allow admin to delete any project, regular users only their own
        if user.is_admin:
            project = Project.query.get(project_id)
            app.logger.info(f"Admin user {user_id_int} attempting to delete project {project_id}")
        else:
            project = Project.query.filter_by(id=project_id, user_id=user_id_int).first()
            app.logger.info(f"Regular user {user_id_int} attempting to delete their project {project_id}")

        if not project:
            app.logger.error(f"Project {project_id} not found or user {user_id_int} doesn't have access")
            return jsonify({"error": "Project not found or access denied"}), 404

        app.logger.info(f"Starting deletion of project {project_id} by user {user_id_int} (admin: {user.is_admin})")

        try:
            # Delete all citations first
            citations_deleted = Citation.query.filter_by(project_id=project_id).delete(synchronize_session=False)
            app.logger.info(f"Deleted {citations_deleted} citations for project {project_id}")

            # Delete the project
            db.session.delete(project)
            db.session.commit()

            app.logger.info(f"Successfully deleted project {project_id}")
            return jsonify({"success": True, "message": "Project deleted successfully"}), 200

        except Exception as delete_error:
            app.logger.error(f"Database error during project deletion: {str(delete_error)}")
            db.session.rollback()
            return jsonify({"error": "Database error during deletion", "details": str(delete_error)}), 500

    except Exception as e:
        app.logger.error(f"General error deleting project {project_id}: {str(e)}")
        db.session.rollback()
        return jsonify({"error": "Failed to delete project", "details": str(e)}), 500

@app.route('/api/projects/<int:project_id>/labeled-citations', methods=['GET'])
def get_labeled_citations(project_id):
    user_id = request.headers.get('X-User-Id')
    if not user_id:
        return jsonify({"error": "Unauthorized"}), 401

    user_id_int = int(user_id)
    user = User.query.get(user_id_int)
    if not user:
        return jsonify({"error": "User not found"}), 401

    # Allow admin access to any project, regular users only their own
    if user.is_admin:
        project = Project.query.get(project_id)
    else:
        project = Project.query.filter_by(id=project_id, user_id=user_id_int).first()

    if not project:
        return jsonify({"error": "Project not found"}), 404

    labeled = Citation.query.filter(Citation.project_id == project_id,
                                    Citation.is_relevant.isnot(None)).all()

    return jsonify({
        "labeled_citations": [{
            "id": c.id,
            "title": c.title,
            "is_relevant": c.is_relevant,
            "iteration": c.iteration
        } for c in labeled]
    })


@app.route('/api/projects/<int:project_id>/remove-duplicates', methods=['POST'])
def remove_duplicates(project_id):
    try:
        user_id = request.headers.get('X-User-Id')
        if not user_id:
            return jsonify({"error": "Unauthorized"}), 401

        user_id_int = int(user_id)
        user = User.query.get(user_id_int)
        if not user:
            return jsonify({"error": "User not found"}), 401

        # Allow admin access to any project, regular users only their own
        if user.is_admin:
            project = Project.query.get(project_id)
        else:
            project = Project.query.filter_by(id=project_id, user_id=user_id_int).first()

        if not project:
            return jsonify({"error": "Project not found"}), 404

        # Get all citations for this project
        citations = Citation.query.filter_by(project_id=project_id).all()

        if not citations:
            return jsonify({"error": "No citations found in this project"}), 404

        app.logger.info(f"Starting advanced duplicate removal for project {project_id} with {len(citations)} citations")

        # Convert citations to DataFrame for easier processing (include metadata)
        citations_data = []
        for i, citation in enumerate(citations):
            row = {
                'id': citation.id,
                'title': citation.title,
                'abstract': citation.abstract,
                'is_relevant': citation.is_relevant,
                'iteration': citation.iteration,
                'index': i
            }
            if getattr(citation, 'extra_metadata', None):
                # Merge metadata keys as columns
                for k, v in citation.extra_metadata.items():
                    # Avoid overwriting core fields
                    if k not in row:
                        row[k] = v
            citations_data.append(row)

        df = pd.DataFrame(citations_data)

        # Apply the enhanced duplicate processing logic with column standardization
        app.logger.info(f"Using enhanced duplicate detection for {len(df)} citations")
        
        # Standardize the DataFrame columns
        df_standardized = standardize_dataframe_columns(df.copy())
        
        # Parse custom config from body (optional)
        body = request.get_json(silent=True) or {}
        mode = (body.get('mode') or (project.duplicate_config or {}).get('mode') or 'default').lower()
        selected_columns = body.get('columns') or (project.duplicate_config or {}).get('columns') or []
        threshold = float(body.get('threshold') or (project.duplicate_config or {}).get('threshold') or 0.9)
        year_field = body.get('year_field') or (project.duplicate_config or {}).get('year_field') or 'publication_year'

        # Persist config on project
        project.duplicate_config = {
            'mode': mode,
            'columns': selected_columns,
            'threshold': threshold,
            'year_field': year_field
        }

        # Check if we have author and year information (default mode path)
        has_authors = 'authors' in df_standardized.columns and not df_standardized['authors'].isna().all()
        has_year = year_field in df_standardized.columns and not df_standardized[year_field].isna().all()
        
        app.logger.info(f"Enhanced duplicate detection: Authors present: {has_authors}, Year present: {has_year}")
        
        # Apply duplicate processing based on mode
        app.logger.info(f"Duplicate mode: {mode}; threshold={threshold}; columns={selected_columns}")
        app.logger.info(f"Calling duplicate processing with DataFrame: {df_standardized.shape}")
        app.logger.info(f"DataFrame columns: {list(df_standardized.columns)}")
        app.logger.info(f"Sample data - First row: {df_standardized.iloc[0].to_dict() if len(df_standardized) > 0 else 'Empty'}")

        if mode == 'custom' and selected_columns:
            result = process_custom_duplicates(df_standardized, selected_columns, threshold, year_field)
            strategy_used = f"Custom TF-IDF on {selected_columns} at {threshold}"
        else:
            result = process_enhanced_duplicates(df_standardized, has_authors, has_year)
            strategy_used = result.get('strategy', 'Enhanced default')
        
        if result['error']:
            app.logger.error(f"Enhanced duplicate processing failed: {result['error']}")
            # Fallback to old advanced processing
            app.logger.info("Falling back to advanced duplicate processing")
            result = process_advanced_duplicates(df)
        else:
            app.logger.info(f"Enhanced duplicate processing successful: {len(result.get('citations_to_keep', []))} citations kept")
            app.logger.info(f"Result keys: {list(result.keys())}")

        if result['error']:
            return jsonify({"error": result['error']}), 400

        # Get citation IDs to keep - handle both enhanced and advanced systems
        if 'kept_indices' in result:
            # Old advanced system response
            kept_indices = result['kept_indices']
            citations_to_keep = [citations[i] for i in kept_indices]
            citations_to_remove = [citations[i] for i in range(len(citations)) if i not in kept_indices]
        else:
            # New enhanced system response
            citations_to_keep = result['citations_to_keep']
            # Find citations to remove by comparing with original citations
            kept_citation_ids = [citation['id'] for citation in citations_to_keep]
            citations_to_remove = [citation for citation in citations if citation.id not in kept_citation_ids]

        # Mark duplicates instead of deleting them
        duplicates_marked = len(citations_to_remove)
        
        if 'kept_indices' in result:
            # Old advanced system - use indices
            citation_ids_to_mark = [citations[i].id for i in range(len(citations)) if i not in kept_indices]
        else:
            # New enhanced system - use citation objects
            citation_ids_to_mark = [citation.id for citation in citations_to_remove]

        if citation_ids_to_mark:
            Citation.query.filter(Citation.id.in_(citation_ids_to_mark)).update(
                {Citation.is_duplicate: True}, synchronize_session=False
            )

        # NEW: Store duplicate details in project record
        project.duplicates_removed = True
        project.duplicates_count = duplicates_marked
        project.citations_count = len(citations_to_keep)
        
        # Store duplicate details based on system used
        if 'kept_indices' in result:
            # Old advanced system
            project.duplicate_details = make_json_serializable(result['duplicate_details'])
            project.processing_summary = make_json_serializable(result['processing_summary'])
            project.removal_strategy = result['removal_strategy']
            app.logger.info(f"Stored old system duplicate details: {len(result['duplicate_details'])} entries")
        else:
            # New enhanced system
            project.duplicate_details = make_json_serializable(result['duplicate_details'])
            project.removal_strategy = strategy_used
            project.detection_method_used = "Custom TF-IDF" if mode == 'custom' and selected_columns else "Enhanced TF-IDF with column standardization"
            project.columns_standardized = True
            
            # Count resolution strategies used
            year_resolution_count = sum(1 for detail in result['duplicate_details'] if detail.get('type') == 'year_resolution')
            abstract_resolution_count = sum(1 for detail in result['duplicate_details'] if detail.get('type') == 'abstract_resolution')
            
            project.year_resolution_count = year_resolution_count
            project.abstract_resolution_count = abstract_resolution_count
            
            app.logger.info(f"Stored enhanced system duplicate details: {len(result['duplicate_details'])} entries")
            app.logger.info(f"Sample duplicate detail: {result['duplicate_details'][0] if result['duplicate_details'] else 'None'}")

        db.session.commit()
        app.logger.info(f"Marked {duplicates_marked} citations as duplicates in project {project_id}")

        # Prepare response based on system used
        if 'kept_indices' in result:
            # Old advanced system response
            response_data = {
                "message": f"Advanced duplicate detection completed",
                "duplicates_marked": duplicates_marked,
                "unique_citations": len(citations_to_keep),
                "removal_strategy": result['removal_strategy'],
                "duplicate_details": result['duplicate_details'][:10],
                "processing_summary": result['processing_summary'],
                
                # Status fields for frontend
                "duplicates_removed": True,
                "duplicates_count": duplicates_marked,
                "citations_count": len(citations_to_keep)
            }
        else:
            # New enhanced system response
            response_data = {
                "message": f"Enhanced duplicate detection completed",
                "duplicates_marked": duplicates_marked,
                "unique_citations": len(citations_to_keep),
                "removal_strategy": strategy_used,
                "duplicate_details": result['duplicate_details'][:10],
                "processing_summary": {
                    "detection_method": "Custom TF-IDF" if mode == 'custom' and selected_columns else "Enhanced TF-IDF with column standardization",
                    "duplicates_found": len(result['duplicate_details']),
                    "strategy_used": strategy_used
                },
                
                # Status fields for frontend
                "duplicates_removed": True,
                "duplicates_count": duplicates_marked,
                "citations_count": len(citations_to_keep),
                "duplicate_config_used": project.duplicate_config
            }
            
            app.logger.info(f"Enhanced system response - duplicate_details count: {len(result['duplicate_details'])}")
            app.logger.info(f"Enhanced system response - sample duplicate_detail: {result['duplicate_details'][0] if result['duplicate_details'] else 'None'}")
        
        return jsonify(response_data), 200

    except Exception as e:
        app.logger.error(f"Error removing duplicates: {str(e)}")
        db.session.rollback()
        return jsonify({
            "error": "Failed to remove duplicates",
            "details": str(e)
        }), 500

def process_advanced_duplicates(df):
    """Advanced duplicate processing with multiple strategies"""

    app.logger.info(f"Processing {len(df)} citations for advanced duplicate detection")

    duplicate_details = []
    processing_summary = {
        'exact_duplicates_found': 0,
        'similar_duplicates_found': 0,
        'hash_groups_processed': 0,
        'tfidf_comparisons': 0
    }

    try:
        # Step 1: Exact duplicate detection using text hashing
        df['text_hash'] = df.apply(lambda row: create_text_hash(row['title'], row['abstract']), axis=1)
        hash_groups = df.groupby('text_hash')
        processing_summary['hash_groups_processed'] = len(hash_groups)

        # Step 2: Process each hash group
        kept_indices = []

        for hash_val, group in hash_groups:
            if len(group) == 1:
                # No duplicates in this group
                kept_indices.append(group.iloc[0]['index'])
            else:
                # Multiple citations with same hash - exact duplicates
                processing_summary['exact_duplicates_found'] += len(group) - 1
                app.logger.info(f"Found {len(group)} exact duplicates with hash {hash_val[:8]}...")

                # Select best citation from exact duplicates
                best_citation = select_best_citation_advanced(group, duplicate_details)
                kept_indices.append(best_citation['index'])

        # Step 3: TF-IDF similarity analysis on remaining citations
        if len(kept_indices) > 1:
            app.logger.info(f"Running TF-IDF similarity analysis on {len(kept_indices)} remaining citations")

            # Get the kept citations for similarity analysis
            remaining_df = df[df['index'].isin(kept_indices)].copy()

            # Calculate TF-IDF similarity matrix
            similarity_result = calculate_tfidf_similarity(remaining_df, duplicate_details)

            if similarity_result['success']:
                # Update kept indices based on similarity analysis
                final_indices = similarity_result['kept_indices']
                processing_summary['similar_duplicates_found'] = len(kept_indices) - len(final_indices)
                processing_summary['tfidf_comparisons'] = similarity_result['comparisons_made']
                kept_indices = final_indices
            else:
                app.logger.warning(f"TF-IDF analysis failed: {similarity_result['error']}")

        removal_strategy = f"Multi-stage: Hash-based exact duplicate detection + TF-IDF similarity analysis (threshold: 0.90)"

        return {
            'error': None,
            'kept_indices': kept_indices,
            'removal_strategy': removal_strategy,
            'duplicate_details': duplicate_details,
            'processing_summary': processing_summary
        }

    except Exception as e:
        app.logger.error(f"Advanced duplicate processing failed: {str(e)}")
        return {
            'error': f"Advanced duplicate processing failed: {str(e)}",
            'kept_indices': list(range(len(df))),
            'removal_strategy': 'Error - no duplicates removed',
            'duplicate_details': [],
            'processing_summary': processing_summary
        }


def select_best_citation_advanced(group, duplicate_details):
    """Select the best citation from a group of exact duplicates using multiple criteria"""

    if len(group) == 1:
        return group.iloc[0]

    # Strategy 1: Prioritize labeled citations (those with relevance feedback)
    labeled_citations = group[group['is_relevant'].notna()]
    if len(labeled_citations) > 0:
        # Among labeled citations, prefer relevant ones
        relevant_citations = labeled_citations[labeled_citations['is_relevant'] == True]
        if len(relevant_citations) > 0:
            best = relevant_citations.iloc[0]
        else:
            best = labeled_citations.iloc[0]

        # Log the selection reason with ENHANCED format
        for _, other in group.iterrows():
            if other['index'] != best['index']:
                detail = create_enhanced_duplicate_detail(
                    best, other, 
                    f'Prioritized labeled citation (relevance: {best["is_relevant"]})'
                )
                duplicate_details.append(detail)
        return best

    # Strategy 2: Use completeness score (abstract length, title quality)
    group_copy = group.copy()
    group_copy['completeness_score'] = group_copy.apply(calculate_advanced_completeness_score, axis=1)
    best = group_copy.loc[group_copy['completeness_score'].idxmax()]

    # Log the selection reason with ENHANCED format
    for _, other in group.iterrows():
        if other['index'] != best['index']:
            detail = create_enhanced_duplicate_detail(
                best, other,
                f'Higher completeness score ({best.get("completeness_score", 0):.2f})'
            )
            duplicate_details.append(detail)

    return best


def calculate_advanced_completeness_score(row):
    """Calculate a comprehensive completeness score for citation quality"""
    score = 0

    # Title quality (10-40 points)
    title_len = len(str(row['title'])) if pd.notna(row['title']) else 0
    if title_len > 100:
        score += 40
    elif title_len > 50:
        score += 30
    elif title_len > 20:
        score += 20
    else:
        score += 10

    # Abstract quality (20-50 points)
    abstract_len = len(str(row['abstract'])) if pd.notna(row['abstract']) else 0
    if abstract_len > 1000:
        score += 50
    elif abstract_len > 500:
        score += 40
    elif abstract_len > 200:
        score += 30
    elif abstract_len > 50:
        score += 20
    else:
        score += 10

    # Bonus for being labeled (10 points)
    if pd.notna(row['is_relevant']):
        score += 10

    return score

def calculate_tfidf_similarity(df, duplicate_details):
    """Calculate TF-IDF similarity and remove similar citations"""

    try:
        # Prepare text for TF-IDF analysis
        texts = []
        for _, row in df.iterrows():
            text = f"{normalize_text(row['title'])} {normalize_text(row['abstract'])}"
            texts.append(text)

        if len(texts) < 2:
            return {'success': True, 'kept_indices': df['index'].tolist(), 'comparisons_made': 0}

        # Calculate TF-IDF vectors
        vectorizer = TfidfVectorizer(
            max_features=1000, 
            stop_words='english', 
            ngram_range=(1, 2),
            min_df=1,
            max_df=0.90
        )

        tfidf_matrix = vectorizer.fit_transform(texts)
        similarity_matrix = cosine_similarity(tfidf_matrix)

        # Find similar pairs using higher threshold (90% similarity)
        similarity_threshold = 0.90
        to_remove = set()
        comparisons_made = 0

        for i in range(len(similarity_matrix)):
            for j in range(i + 1, len(similarity_matrix)):
                comparisons_made += 1
                similarity_score = similarity_matrix[i][j]

                if similarity_score > similarity_threshold:
                    row_i = df.iloc[i]
                    row_j = df.iloc[j]

                    # Decide which one to keep based on multiple criteria
                    keep_i = should_keep_citation_i(row_i, row_j)

                    if keep_i:
                        to_remove.add(j)
                        # Use ENHANCED format with full information
                        detail = create_enhanced_duplicate_detail(
                            row_i, row_j,
                            f'TF-IDF similarity above threshold',
                            similarity_score
                        )
                        duplicate_details.append(detail)
                    else:
                        to_remove.add(i)
                        # Use ENHANCED format with full information
                        detail = create_enhanced_duplicate_detail(
                            row_j, row_i,
                            f'TF-IDF similarity above threshold',
                            similarity_score
                        )
                        duplicate_details.append(detail)

        # Return indices of citations to keep
        kept_indices = [df.iloc[i]['index'] for i in range(len(df)) if i not in to_remove]

        return {
            'success': True,
            'kept_indices': kept_indices,
            'comparisons_made': comparisons_made
        }

    except Exception as e:
        app.logger.warning(f"TF-IDF similarity calculation failed: {str(e)}")
        return {
            'success': False,
            'error': str(e),
            'kept_indices': df['index'].tolist(),
            'comparisons_made': 0
        }



def should_keep_citation_i(citation_i, citation_j):
    """Decide which citation to keep when they are similar"""

    # Priority 1: Keep labeled citations
    i_labeled = pd.notna(citation_i['is_relevant'])
    j_labeled = pd.notna(citation_j['is_relevant'])

    if i_labeled and not j_labeled:
        return True
    elif j_labeled and not i_labeled:
        return False
    elif i_labeled and j_labeled:
        # Both labeled - prefer relevant ones
        if citation_i['is_relevant'] == True and citation_j['is_relevant'] != True:
            return True
        elif citation_j['is_relevant'] == True and citation_i['is_relevant'] != True:
            return False

    # Priority 2: Compare completeness scores
    score_i = calculate_advanced_completeness_score(citation_i)
    score_j = calculate_advanced_completeness_score(citation_j)

    return score_i >= score_j

def process_enhanced_duplicates(df, has_authors, has_year):
    """Main function to process duplicates using enhanced logic"""
    try:
        app.logger.info(f"Starting enhanced duplicate processing. Authors: {has_authors}, Year: {has_year}")
        app.logger.info(f"Input DataFrame shape: {df.shape}")
        app.logger.info(f"Input DataFrame columns: {list(df.columns)}")
        
        # Validate required columns
        if 'title' not in df.columns:
            app.logger.error(f"Title column not found. Available columns: {list(df.columns)}")
            return {'error': 'Title column is required but not found', 'citations_to_keep': [], 'duplicate_details': [], 'strategy': 'Error'}
        
        if 'abstract' not in df.columns:
            app.logger.error(f"Abstract column not found. Available columns: {list(df.columns)}")
            return {'error': 'Abstract column is required but not found', 'citations_to_keep': [], 'duplicate_details': [], 'strategy': 'Error'}
        
        duplicate_details = []
        citations_to_keep = []
        
        if has_authors:
            # Scenario 1: Authors present - check title + authors
            app.logger.info("Processing duplicates with author information")
            result = detect_duplicates_with_authors(df, has_year)
            citations_to_keep = result['citations_to_keep']
            duplicate_details.extend(result['duplicate_details'])
            strategy = "Title + Authors with 90% TF-IDF similarity"
        else:
            # Scenario 2: No authors - check title only
            app.logger.info("Processing duplicates with title only")
            result = detect_duplicates_title_only(df, has_year)
            citations_to_keep = result['citations_to_keep']
            duplicate_details.extend(result['duplicate_details'])
            strategy = "Title only with 90% TF-IDF similarity"
        
        app.logger.info(f"Enhanced duplicate processing completed. Keeping {len(citations_to_keep)} citations")
        
        return {
            'error': None,
            'citations_to_keep': citations_to_keep,
            'duplicate_details': duplicate_details,
            'strategy': strategy
        }
        
    except Exception as e:
        app.logger.error(f"Error in enhanced duplicate processing: {str(e)}")
        return {
            'error': f'Enhanced duplicate processing failed: {str(e)}',
            'citations_to_keep': [],
            'duplicate_details': [],
            'strategy': 'Error'
        }

def detect_duplicates_with_authors(df, has_year):
    """Detect duplicates when authors are present (Title + Authors)"""
    duplicate_details = []
    citations_to_keep = []
    
    # Create combined text for comparison (title + authors)
    df['combined_text'] = df['title'].astype(str) + ' ' + df['authors'].astype(str)
    
    # Apply TF-IDF similarity detection
    vectorizer = TfidfVectorizer(
        max_features=1000,
        stop_words='english',
        ngram_range=(1, 2),
        min_df=1
    )
    
    try:
        tfidf_matrix = vectorizer.fit_transform(df['combined_text'].fillna(''))
        similarity_matrix = cosine_similarity(tfidf_matrix)
        
        # Find groups of similar citations (90% threshold)
        processed_indices = set()
        
        for i in range(len(df)):
            if i in processed_indices:
                continue
                
            # Find all citations similar to this one
            similar_indices = [i]
            for j in range(i + 1, len(df)):
                if j not in processed_indices and similarity_matrix[i][j] >= 0.9:
                    similar_indices.append(j)
            
            if len(similar_indices) > 1:
                # Duplicates found - resolve using year or abstract length
                group = df.iloc[similar_indices]
                best_citation = resolve_duplicates_by_strategy(group, has_year, duplicate_details)
                citations_to_keep.append(best_citation)
                
                # Mark all as processed
                processed_indices.update(similar_indices)
                
                app.logger.info(f"Found {len(similar_indices)} duplicates, kept 1 based on strategy")
            else:
                # No duplicates - keep this citation
                citations_to_keep.append(df.iloc[i])
                processed_indices.add(i)
        
        return {
            'citations_to_keep': citations_to_keep,
            'duplicate_details': duplicate_details
        }
        
    except Exception as e:
        app.logger.error(f"Error in author-based duplicate detection: {str(e)}")
        # Fallback: return all citations
        return {
            'citations_to_keep': [df.iloc[i] for i in range(len(df))],
            'duplicate_details': [{'error': f'Author-based detection failed: {str(e)}'}]
        }

def detect_duplicates_title_only(df, has_year):
    """Detect duplicates when authors are not present (Title only)"""
    app.logger.info(f"Starting title-only duplicate detection. DataFrame shape: {df.shape}")
    app.logger.info(f"Sample titles: {list(df['title'].head(3))}")
    
    duplicate_details = []
    citations_to_keep = []
    
    # Apply TF-IDF similarity detection on title only
    vectorizer = TfidfVectorizer(
        max_features=1000,
        stop_words='english',
        ngram_range=(1, 2),
        min_df=1
    )
    
    try:
        tfidf_matrix = vectorizer.fit_transform(df['title'].fillna(''))
        similarity_matrix = cosine_similarity(tfidf_matrix)
        
        # Debug: Print similarity matrix
        app.logger.info(f"DEBUG: TF-IDF matrix shape: {tfidf_matrix.shape}")
        app.logger.info(f"DEBUG: Similarity matrix shape: {similarity_matrix.shape}")
        for i in range(len(similarity_matrix)):
            for j in range(i + 1, len(similarity_matrix)):
                app.logger.info(f"DEBUG: Similarity [{i}][{j}] = {similarity_matrix[i][j]:.3f}")
        
        # Find groups of similar citations (80% threshold for title-only)
        processed_indices = set()
        
        for i in range(len(df)):
            if i in processed_indices:
                continue
                
            # Find all citations similar to this one
            similar_indices = [i]
            for j in range(i + 1, len(df)):
                # Use 90% threshold for title-only comparison
                similarity_threshold = 0.9  # 90% threshold
                if j not in processed_indices and similarity_matrix[i][j] >= similarity_threshold:
                    app.logger.info(f"DEBUG: Found duplicate with similarity {similarity_matrix[i][j]:.3f} >= {similarity_threshold}")
                    similar_indices.append(j)
                else:
                    app.logger.info(f"DEBUG: Not a duplicate: similarity {similarity_matrix[i][j]:.3f} < {similarity_threshold}")
            
            if len(similar_indices) > 1:
                # Duplicates found - resolve using year or abstract length
                group = df.iloc[similar_indices]
                best_citation = resolve_duplicates_by_strategy(group, has_year, duplicate_details)
                citations_to_keep.append(best_citation)
                
                # Mark all as processed
                processed_indices.update(similar_indices)
                
                app.logger.info(f"Found {len(similar_indices)} duplicates, kept 1 based on strategy")
            else:
                # No duplicates - keep this citation
                citations_to_keep.append(df.iloc[i])
                processed_indices.add(i)
        
        return {
            'citations_to_keep': citations_to_keep,
            'duplicate_details': duplicate_details
        }
        
    except Exception as e:
        app.logger.error(f"Error in title-only duplicate detection: {str(e)}")
        # Fallback: return all citations
        return {
            'citations_to_keep': [df.iloc[i] for i in range(len(df))],
            'duplicate_details': [{'error': f'Title-only detection failed: {str(e)}'}]
        }

def resolve_duplicates_by_strategy(group, has_year, duplicate_details):
    """Resolve duplicates using year or abstract length strategy"""
    if has_year:
        # Strategy 1: Keep the one with older published year
        year_resolution_count = 0
        best_citation = None
        oldest_year = float('inf')
        
        for _, row in group.iterrows():
            year = extract_year_from_value(row.get('publication_year'))
            if year and year < oldest_year:
                oldest_year = year
                best_citation = row
                year_resolution_count += 1
        
        if best_citation is not None:
            # Store detailed information about kept vs removed citations
            for _, row in group.iterrows():
                if row.name != best_citation.name:
                    duplicate_details.append({
                        'type': 'year_resolution',
                        'kept': {
                            'id': int(best_citation.get('id', 0)),
                            'title': str(best_citation.get('title', 'N/A')),
                            'abstract': str(best_citation.get('abstract', 'N/A'))[:100] + '...' if len(str(best_citation.get('abstract', ''))) > 100 else str(best_citation.get('abstract', 'N/A'))
                        },
                        'removed': {
                            'id': int(row.get('id', 0)),
                            'title': str(row.get('title', 'N/A')),
                            'abstract': str(row.get('abstract', 'N/A'))[:100] + '...' if len(str(row.get('abstract', ''))) > 100 else str(row.get('abstract', 'N/A'))
                        },
                        'reason': f'Older publication year ({oldest_year})',
                        'strategy': 'Year-based resolution'
                    })
            return best_citation
    
    # Strategy 2: Keep the one with longer abstract content
    abstract_resolution_count = 0
    best_citation = None
    max_abstract_length = 0
    
    for _, row in group.iterrows():
        abstract = str(row.get('abstract', ''))
        if len(abstract) > max_abstract_length:
            max_abstract_length = len(abstract)
            best_citation = row
            abstract_resolution_count += 1
    
    if best_citation is not None:
        # Store detailed information about kept vs removed citations
        for _, row in group.iterrows():
            if row.name != best_citation.name:
                duplicate_details.append({
                    'type': 'abstract_resolution',
                    'kept': {
                        'id': int(best_citation.get('id', 0)),
                        'title': str(best_citation.get('title', 'N/A')),
                        'abstract': str(best_citation.get('abstract', 'N/A'))[:100] + '...' if len(str(best_citation.get('abstract', ''))) > 100 else str(best_citation.get('abstract', 'N/A'))
                    },
                    'removed': {
                        'id': int(row.get('id', 0)),
                        'title': str(row.get('title', 'N/A')),
                        'abstract': str(row.get('abstract', 'N/A'))[:100] + '...' if len(str(row.get('abstract', ''))) > 100 else str(row.get('abstract', 'N/A'))
                    },
                    'reason': f'Longer abstract content ({max_abstract_length} characters)',
                    'strategy': 'Abstract length resolution'
                })
        return best_citation
    
    # If all else fails, return the first citation
    if best_citation is None:
        best_citation = group.iloc[0]
        # Store detailed information about kept vs removed citations
        for _, row in group.iterrows():
            if row.name != best_citation.name:
                duplicate_details.append({
                    'type': 'fallback',
                    'kept': {
                        'id': int(best_citation.get('id', 0)),
                        'title': str(best_citation.get('title', 'N/A')),
                        'abstract': str(best_citation.get('abstract', 'N/A'))[:100] + '...' if len(str(best_citation.get('abstract', ''))) > 100 else str(best_citation.get('abstract', 'N/A'))
                    },
                    'removed': {
                        'id': int(row.get('id', 0)),
                        'title': str(row.get('title', 'N/A')),
                        'abstract': str(row.get('abstract', 'N/A'))[:100] + '...' if len(str(row.get('abstract', ''))) > 100 else str(best_citation.get('abstract', 'N/A'))
                    },
                    'reason': 'First citation (fallback)',
                    'strategy': 'Fallback resolution'
                })
    
    return best_citation

def standardize_dataframe_columns(df):
    """Standardize column names to consistent format for duplicate detection"""
    df_normalized = df.copy()
    
    # Standardize column names to lowercase and remove extra spaces
    df_normalized.columns = df_normalized.columns.str.lower().str.strip()
    
    # Map common column name variations to standard names
    column_mapping = {}
    
    # Year column variations
    year_patterns = ['year', 'publication_year', 'pub_year', 'published_year', 
                    'publication_date', 'pub_date', 'date', 'published', 'pubdate']
    year_col = find_column_by_patterns(df_normalized, year_patterns)
    if year_col:
        column_mapping[year_col] = 'publication_year'
    
    # Author column variations
    author_patterns = ['author', 'authors', 'published_author', 'published_authors', 
                      'by', 'written_by', 'creator', 'creators']
    author_col = find_column_by_patterns(df_normalized, author_patterns)
    if author_col:
        column_mapping[author_col] = 'authors'
    
    # Title column variations
    title_patterns = ['title', 'name', 'heading', 'headline', 'citation_title']
    title_col = find_column_by_patterns(df_normalized, title_patterns)
    if title_col:
        column_mapping[title_col] = 'title'
    
    # Abstract column variations
    abstract_patterns = ['abstract', 'summary', 'description', 'content', 'text', 'body']
    abstract_col = find_column_by_patterns(df_normalized, abstract_patterns)
    if abstract_col:
        column_mapping[abstract_col] = 'abstract'
    
    # Rename columns to standard format
    df_normalized = df_normalized.rename(columns=column_mapping)
    
    app.logger.info(f"Column standardization completed. Mapped columns: {column_mapping}")
    return df_normalized

def find_column_by_patterns(df, patterns):
    """Find column by matching patterns (case insensitive)"""
    for col in df.columns:
        for pattern in patterns:
            if pattern.lower() in col.lower():
                return col
    return None

# Helpers for custom duplicate detection
def build_similarity_text(row, selected_columns):
    parts = []
    for col in selected_columns:
        value = row.get(col)
        if pd.isna(value):
            continue
        parts.append(str(value))
    return ' '.join(parts)

def process_custom_duplicates(df, selected_columns, threshold, year_field):
    """Detect duplicates using custom-selected columns for TF-IDF similarity."""
    duplicate_details = []
    citations_to_keep = []

    # Ensure selected columns exist; quietly ignore missing
    cols = [c for c in selected_columns if c in df.columns]
    if not cols:
        return {
            'citations_to_keep': [df.iloc[i] for i in range(len(df))],
            'duplicate_details': [{'error': 'No valid columns provided for custom detection'}],
            'strategy': 'Custom columns - no-op'
        }

    df = df.copy()
    df['combined_text'] = df.apply(lambda r: build_similarity_text(r, cols), axis=1)

    vectorizer = TfidfVectorizer(
        max_features=1000,
        stop_words='english',
        ngram_range=(1, 2),
        min_df=1
    )

    try:
        tfidf_matrix = vectorizer.fit_transform(df['combined_text'].fillna(''))
        similarity_matrix = cosine_similarity(tfidf_matrix)

        processed_indices = set()
        for i in range(len(df)):
            if i in processed_indices:
                continue
            similar_indices = [i]
            for j in range(i + 1, len(df)):
                if j not in processed_indices and similarity_matrix[i][j] >= float(threshold):
                    similar_indices.append(j)

            if len(similar_indices) > 1:
                group = df.iloc[similar_indices]
                has_year = (year_field in df.columns) and (not df[year_field].isna().all())
                best_citation = resolve_duplicates_by_strategy(group, has_year, duplicate_details)
                citations_to_keep.append(best_citation)
                processed_indices.update(similar_indices)
            else:
                citations_to_keep.append(df.iloc[i])
                processed_indices.add(i)

        return {
            'citations_to_keep': citations_to_keep,
            'duplicate_details': duplicate_details,
            'strategy': f'Custom TF-IDF on {cols} at {threshold}'
        }
    except Exception as e:
        app.logger.error(f"Error in custom duplicate detection: {str(e)}")
        return {
            'citations_to_keep': [df.iloc[i] for i in range(len(df))],
            'duplicate_details': [{'error': f'Custom detection failed: {str(e)}'}],
            'strategy': 'Custom - error fallback'
        }

# At the end of your main.py file:
if __name__ == "__main__":
    port = int(os.environ.get('PORT',5000))
    app.run(host='0.0.0.0', port=port,debug=False)
